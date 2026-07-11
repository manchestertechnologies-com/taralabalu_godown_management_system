import os
import sqlite3
import urllib.parse
import psycopg2
from datetime import datetime


# Supabase Connection String (Default)
SUPABASE_URL = "postgresql://postgres:Bery8792480218@db.rsvkrseaxewamlipnuku.supabase.co:5432/postgres"
SUPABASE_URL_ALT = "postgresql://postgres:[Bery8792480218]@db.rsvkrseaxewamlipnuku.supabase.co:5432/postgres"

def get_connection():
    # Try connecting to PostgreSQL / Supabase
    connection_strings = [
        os.environ.get('DATABASE_URL', SUPABASE_URL),
        SUPABASE_URL_ALT
    ]
    
    for conn_str in connection_strings:
        # Clean up any potential brackets in connection URL
        cleaned_str = conn_str.replace(":[Bery8792480218]@", ":Bery8792480218@")
        try:
            print(f"Attempting to connect to PostgreSQL (Supabase)...")
            # Parse URL into parts to connect via parameters (handles odd characters safely)
            result = urllib.parse.urlparse(cleaned_str)
            username = result.username
            password = result.password
            database = result.path[1:]
            hostname = result.hostname
            port = result.port
            
            conn = psycopg2.connect(
                database=database,
                user=username,
                password=password,
                host=hostname,
                port=port,
                connect_timeout=3
            )
            print("Successfully connected to PostgreSQL (Supabase).")
            return conn, "postgresql"
        except Exception as e:
            print(f"PostgreSQL connection failed with URL: {cleaned_str.split('@')[-1]}. Error: {e}")
            
    # Fallback to local SQLite using absolute path
    base_dir = os.path.dirname(os.path.abspath(__file__))
    db_path = os.path.join(base_dir, 'database.db')
    print(f"Falling back to local SQLite database '{db_path}'...")
    conn = sqlite3.connect(db_path)
    return conn, "sqlite"


def execute_ddl(cursor, db_type, query_pg, query_sqlite):
    query = query_pg if db_type == "postgresql" else query_sqlite
    cursor.execute(query)

def init_db():
    conn, db_type = get_connection()
    cursor = conn.cursor()
    
    # 1. Institutions Table
    if db_type == "postgresql":
        cursor.execute("DROP TABLE IF EXISTS Institutions CASCADE;")
        cursor.execute("""
            CREATE TABLE Institutions (
                Inst_ID SERIAL PRIMARY KEY,
                Institution VARCHAR(100) UNIQUE NOT NULL
            );
        """)
    else:
        cursor.execute("DROP TABLE IF EXISTS Institutions;")
        cursor.execute("""
            CREATE TABLE Institutions (
                Inst_ID INTEGER PRIMARY KEY AUTOINCREMENT,
                Institution TEXT UNIQUE NOT NULL
            );
        """)

    # 2. Shops_Donors Table
    if db_type == "postgresql":
        cursor.execute("DROP TABLE IF EXISTS Shops_Donors CASCADE;")
        cursor.execute("""
            CREATE TABLE Shops_Donors (
                Rec SERIAL PRIMARY KEY,
                Year1 VARCHAR(10),
                Shop_Donor_ID INTEGER UNIQUE NOT NULL,
                Shop_Donor_Name VARCHAR(200) NOT NULL,
                Place_ID INTEGER,
                Place VARCHAR(100),
                Place_Kan VARCHAR(100),
                Taluk VARCHAR(100),
                Taluk_Kan VARCHAR(100),
                Dist VARCHAR(100),
                Dist_Kan VARCHAR(100),
                Pin VARCHAR(20),
                State VARCHAR(100),
                Country VARCHAR(100),
                Mobile VARCHAR(20),
                Remarks TEXT,
                DateStamp VARCHAR(50)
            );
        """)
    else:
        cursor.execute("DROP TABLE IF EXISTS Shops_Donors;")
        cursor.execute("""
            CREATE TABLE Shops_Donors (
                Rec INTEGER PRIMARY KEY AUTOINCREMENT,
                Year1 TEXT,
                Shop_Donor_ID INTEGER UNIQUE NOT NULL,
                Shop_Donor_Name TEXT NOT NULL,
                Place_ID INTEGER,
                Place TEXT,
                Place_Kan TEXT,
                Taluk TEXT,
                Taluk_Kan TEXT,
                Dist TEXT,
                Dist_Kan TEXT,
                Pin TEXT,
                State TEXT,
                Country TEXT,
                Mobile TEXT,
                Remarks TEXT,
                DateStamp TEXT
            );
        """)

    # 3. Grocery_Items Table
    if db_type == "postgresql":
        cursor.execute("DROP TABLE IF EXISTS Grocery_Items CASCADE;")
        cursor.execute("""
            CREATE TABLE Grocery_Items (
                Rec SERIAL PRIMARY KEY,
                Grocery_Code INTEGER UNIQUE NOT NULL,
                Grocery_Items_Kan VARCHAR(200) NOT NULL,
                Grocery_Items_Eng VARCHAR(200),
                Grocery_Category VARCHAR(100),
                Category_Code VARCHAR(20),
                Std_Rate REAL DEFAULT 0.0,
                Qtl_Kg_Ltr VARCHAR(20),
                Shraddanjali_Qty REAL DEFAULT 0.0,
                Hunnime_Qty REAL DEFAULT 0.0,
                Boys_Hostel_Qty REAL DEFAULT 0.0,
                Girls_Hostel_Qty REAL DEFAULT 0.0,
                Math_Qty REAL DEFAULT 0.0,
                Shantivan_Qty_a REAL DEFAULT 0.0,
                AO_Office_Qty REAL DEFAULT 0.0,
                Shraddanjali_Budget REAL DEFAULT 0.0,
                Hunnime_Budget REAL DEFAULT 0.0,
                Boys_Hostel_Budget REAL DEFAULT 0.0,
                Girls_Hostel_Budget REAL DEFAULT 0.0,
                Math_Budget REAL DEFAULT 0.0,
                Shantivan_Budget_a REAL DEFAULT 0.0,
                AO_Office_Budget REAL DEFAULT 0.0,
                Tot_Quantity REAL DEFAULT 0.0,
                Opening_Stock REAL DEFAULT 0.0,
                Closing_Stock REAL DEFAULT 0.0,
                Tot_Stock REAL DEFAULT 0.0,
                Tot_Issue REAL DEFAULT 0.0,
                Stock_Amt REAL DEFAULT 0.0,
                Total_Budget REAL DEFAULT 0.0,
                Remarks TEXT,
                DateStamp VARCHAR(50)
            );
        """)
    else:
        cursor.execute("DROP TABLE IF EXISTS Grocery_Items;")
        cursor.execute("""
            CREATE TABLE Grocery_Items (
                Rec INTEGER PRIMARY KEY AUTOINCREMENT,
                Grocery_Code INTEGER UNIQUE NOT NULL,
                Grocery_Items_Kan TEXT NOT NULL,
                Grocery_Items_Eng TEXT,
                Grocery_Category TEXT,
                Category_Code TEXT,
                Std_Rate REAL DEFAULT 0.0,
                Qtl_Kg_Ltr TEXT,
                Shraddanjali_Qty REAL DEFAULT 0.0,
                Hunnime_Qty REAL DEFAULT 0.0,
                Boys_Hostel_Qty REAL DEFAULT 0.0,
                Girls_Hostel_Qty REAL DEFAULT 0.0,
                Math_Qty REAL DEFAULT 0.0,
                Shantivan_Qty_a REAL DEFAULT 0.0,
                AO_Office_Qty REAL DEFAULT 0.0,
                Shraddanjali_Budget REAL DEFAULT 0.0,
                Hunnime_Budget REAL DEFAULT 0.0,
                Boys_Hostel_Budget REAL DEFAULT 0.0,
                Girls_Hostel_Budget REAL DEFAULT 0.0,
                Math_Budget REAL DEFAULT 0.0,
                Shantivan_Budget_a REAL DEFAULT 0.0,
                AO_Office_Budget REAL DEFAULT 0.0,
                Tot_Quantity REAL DEFAULT 0.0,
                Opening_Stock REAL DEFAULT 0.0,
                Closing_Stock REAL DEFAULT 0.0,
                Tot_Stock REAL DEFAULT 0.0,
                Tot_Issue REAL DEFAULT 0.0,
                Stock_Amt REAL DEFAULT 0.0,
                Total_Budget REAL DEFAULT 0.0,
                Remarks TEXT,
                DateStamp TEXT
            );
        """)

    # 4. Stock_Issue Table
    if db_type == "postgresql":
        cursor.execute("DROP TABLE IF EXISTS Stock_Issue CASCADE;")
        cursor.execute("""
            CREATE TABLE Stock_Issue (
                Rec SERIAL PRIMARY KEY,
                Year1 VARCHAR(10),
                Date1 VARCHAR(20),
                Grocery_Code INTEGER REFERENCES Grocery_Items(Grocery_Code),
                Shop_Donor_ID INTEGER,
                Book_No VARCHAR(50),
                Receipt_No VARCHAR(50),
                Receipt_Date VARCHAR(20),
                Bill_Date VARCHAR(20),
                Bill_No VARCHAR(50),
                Purchase_Rate REAL DEFAULT 0.0,
                Purchase_Amount REAL DEFAULT 0.0,
                Paid VARCHAR(20),
                Stock REAL DEFAULT 0.0,
                Issue_Inst_ID INTEGER REFERENCES Institutions(Inst_ID),
                Issue REAL DEFAULT 0.0,
                Issue_Amount REAL DEFAULT 0.0,
                Received_By VARCHAR(100),
                DateStamp VARCHAR(50),
                Remarks TEXT,
                Purchased_Donation VARCHAR(20)
            );
        """)
    else:
        cursor.execute("DROP TABLE IF EXISTS Stock_Issue;")
        cursor.execute("""
            CREATE TABLE Stock_Issue (
                Rec INTEGER PRIMARY KEY AUTOINCREMENT,
                Year1 TEXT,
                Date1 TEXT,
                Grocery_Code INTEGER,
                Shop_Donor_ID INTEGER,
                Book_No TEXT,
                Receipt_No TEXT,
                Receipt_Date TEXT,
                Bill_Date TEXT,
                Bill_No TEXT,
                Purchase_Rate REAL DEFAULT 0.0,
                Purchase_Amount REAL DEFAULT 0.0,
                Paid TEXT,
                Stock REAL DEFAULT 0.0,
                Issue_Inst_ID INTEGER,
                Issue REAL DEFAULT 0.0,
                Issue_Amount REAL DEFAULT 0.0,
                Received_By TEXT,
                DateStamp TEXT,
                Remarks TEXT,
                Purchased_Donation TEXT,
                FOREIGN KEY(Grocery_Code) REFERENCES Grocery_Items(Grocery_Code),
                FOREIGN KEY(Issue_Inst_ID) REFERENCES Institutions(Inst_ID)
            );
        """)

    # 5. Bills Table
    if db_type == "postgresql":
        cursor.execute("DROP TABLE IF EXISTS Bills CASCADE;")
        cursor.execute("""
            CREATE TABLE Bills (
                Rec SERIAL PRIMARY KEY,
                Year1 VARCHAR(10),
                Shop_Donor_ID INTEGER,
                Bill_Date VARCHAR(20),
                Bill_No VARCHAR(50),
                Bill_Amount REAL DEFAULT 0.0,
                Paid_By VARCHAR(100),
                Ch_Date VARCHAR(20),
                Ch_No VARCHAR(50),
                Ch_Amount REAL DEFAULT 0.0,
                Remarks TEXT,
                DateAdded VARCHAR(50)
            );
        """)
    else:
        cursor.execute("DROP TABLE IF EXISTS Bills;")
        cursor.execute("""
            CREATE TABLE Bills (
                Rec INTEGER PRIMARY KEY AUTOINCREMENT,
                Year1 TEXT,
                Shop_Donor_ID INTEGER,
                Bill_Date TEXT,
                Bill_No TEXT,
                Bill_Amount REAL DEFAULT 0.0,
                Paid_By TEXT,
                Ch_Date TEXT,
                Ch_No TEXT,
                Ch_Amount REAL DEFAULT 0.0,
                Remarks TEXT,
                DateAdded TEXT
            );
        """)

    # 6. Vegetable Table
    if db_type == "postgresql":
        cursor.execute("DROP TABLE IF EXISTS Vegetable CASCADE;")
        cursor.execute("""
            CREATE TABLE Vegetable (
                Rec SERIAL PRIMARY KEY,
                Inst_ID INTEGER REFERENCES Institutions(Inst_ID),
                Year1 VARCHAR(10),
                Purchase_On VARCHAR(20),
                V_Code VARCHAR(20),
                Quantity REAL DEFAULT 0.0,
                Bill_Date VARCHAR(20),
                Bill_No VARCHAR(50),
                Rate REAL DEFAULT 0.0,
                Remarks TEXT,
                Issue_Place VARCHAR(100),
                Purchased_Donation VARCHAR(20)
            );
        """)
    else:
        cursor.execute("DROP TABLE IF EXISTS Vegetable;")
        cursor.execute("""
            CREATE TABLE Vegetable (
                Rec INTEGER PRIMARY KEY AUTOINCREMENT,
                Inst_ID INTEGER,
                Year1 TEXT,
                Purchase_On TEXT,
                V_Code TEXT,
                Quantity REAL DEFAULT 0.0,
                Bill_Date TEXT,
                Bill_No TEXT,
                Rate REAL DEFAULT 0.0,
                Remarks TEXT,
                Issue_Place TEXT,
                Purchased_Donation TEXT,
                FOREIGN KEY(Inst_ID) REFERENCES Institutions(Inst_ID)
            );
        """)

    # 7. Indents Table
    if db_type == "postgresql":
        cursor.execute("DROP TABLE IF EXISTS Indents CASCADE;")
        cursor.execute("""
            CREATE TABLE Indents (
                Rec SERIAL PRIMARY KEY,
                Year1 VARCHAR(10),
                Indent_Date VARCHAR(20),
                Grocery_Code INTEGER REFERENCES Grocery_Items(Grocery_Code),
                Inst_ID INTEGER REFERENCES Institutions(Inst_ID),
                Quantity REAL DEFAULT 0.0,
                Indent_no VARCHAR(50),
                Sanctioned_Quantity REAL DEFAULT 0.0,
                Sanctioned VARCHAR(50) DEFAULT 'Pending', -- 'Pending', 'Sent', 'Received'
                Sanctioned_on VARCHAR(20),
                DateStamp VARCHAR(50),
                Remarks TEXT
            );
        """)
    else:
        cursor.execute("DROP TABLE IF EXISTS Indents;")
        cursor.execute("""
            CREATE TABLE Indents (
                Rec INTEGER PRIMARY KEY AUTOINCREMENT,
                Year1 TEXT,
                Indent_Date TEXT,
                Grocery_Code INTEGER,
                Inst_ID INTEGER,
                Quantity REAL DEFAULT 0.0,
                Indent_no TEXT,
                Sanctioned_Quantity REAL DEFAULT 0.0,
                Sanctioned TEXT DEFAULT 'Pending',
                Sanctioned_on TEXT,
                DateStamp TEXT,
                Remarks TEXT,
                FOREIGN KEY(Grocery_Code) REFERENCES Grocery_Items(Grocery_Code),
                FOREIGN KEY(Inst_ID) REFERENCES Institutions(Inst_ID)
            );
        """)

    # --- SEEDING PRE-FILL DATA ---

    # Institutions Seed
    institutions = [
        ("Stores - Boys Hostel",),
        ("Stores - Girls Hostel",),
        ("Stores - Math",),
        ("Stores - Shantivana Bidara",),
        ("Stores - Shantivana Gurukula",),
        ("Stores - Sirigere BHS",),
        ("Stores - Sirigere GHS",),
        ("Stores - A",),
        ("Stores - AO_Office",),
        ("Store - Shraddhajali",)
    ]
    cursor.executemany("INSERT INTO Institutions (Institution) VALUES (%s)" if db_type == "postgresql" else "INSERT INTO Institutions (Institution) VALUES (?)", institutions)


    # Shops / Donors Seed - Load from Excel Register
    import os as _os, glob as _glob
    _base_dir = _os.path.dirname(_os.path.abspath(__file__))
    _donors_loaded = False
    _year_now = str(datetime.now().year)
    _today = datetime.now().strftime('%Y-%m-%d')
    
    # Try all xlsx files for a Shops_Donors sheet
    _xlsx_candidates = [f for f in _os.listdir(_base_dir) if f.endswith('.xlsx') and not f.startswith('~$')]
    for _xf in sorted(_xlsx_candidates):
        try:
            import openpyxl as _oxl
            _wb = _oxl.load_workbook(_os.path.join(_base_dir, _xf), data_only=True)
            if 'Shops_Donors' not in _wb.sheetnames:
                continue
            _ws = _wb['Shops_Donors']
            _rows = list(_ws.iter_rows(values_only=True))
            if len(_rows) < 2:
                continue
            _headers = [str(h).strip().lower() if h else '' for h in _rows[0]]
            def _col(name, alt=None):
                for n in ([name] + ([alt] if alt else [])):
                    if n in _headers:
                        return _headers.index(n)
                return None
            _id_col   = _col('shop_donor_id')
            _name_col = _col('name', 'shop_donor_name')
            _pid_col  = _col('place_id')
            _place_col= _col('place')
            _taluk_col= _col('taluk')
            _dist_col = _col('dist', 'district')
            _pin_col  = _col('pin')
            _mob_col  = _col('mobile')
            _rem_col  = _col('remarks')
            if _id_col is None or _name_col is None:
                continue
            _donors = []
            for _r in _rows[1:]:
                if not _r or all(c is None for c in _r):
                    continue
                try:
                    _did = int(float(_r[_id_col]))
                except:
                    continue
                _dname = str(_r[_name_col] or '').strip()
                if not _dname:
                    continue
                _pid   = int(float(_r[_pid_col])) if _pid_col is not None and _r[_pid_col] else 0
                _place = str(_r[_place_col] or '').strip() if _place_col is not None else ''
                _taluk = str(_r[_taluk_col] or '').strip() if _taluk_col is not None else ''
                _dist  = str(_r[_dist_col] or '').strip() if _dist_col is not None else ''
                _pin   = str(int(float(_r[_pin_col]))) if _pin_col is not None and _r[_pin_col] else ''
                _mob   = str(_r[_mob_col] or '').strip() if _mob_col is not None else ''
                _rem   = str(_r[_rem_col] or '').strip() if _rem_col is not None else ''
                _donors.append((_year_now, _did, _dname, _pid, _place, _place, _taluk, _taluk, _dist, _dist, _pin, 'Karnataka', 'India', _mob, _rem, _today))
            if _donors:
                cursor.executemany("""
                    INSERT INTO Shops_Donors 
                    (Year1, Shop_Donor_ID, Shop_Donor_Name, Place_ID, Place, Place_Kan, Taluk, Taluk_Kan, Dist, Dist_Kan, Pin, State, Country, Mobile, Remarks, DateStamp)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """ if db_type == "postgresql" else """
                    INSERT INTO Shops_Donors 
                    (Year1, Shop_Donor_ID, Shop_Donor_Name, Place_ID, Place, Place_Kan, Taluk, Taluk_Kan, Dist, Dist_Kan, Pin, State, Country, Mobile, Remarks, DateStamp)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, _donors)
                print(f"Loaded {len(_donors)} donors from '{_xf}' sheet 'Shops_Donors'.")
                _donors_loaded = True
                break
        except Exception as _de:
            print(f"Donors load error from {_xf}: {_de}")
    
    if not _donors_loaded:
        # Fallback minimal seed
        _fallback = [(_year_now, 101, 'Davangere Nagara Baktadigalu', 0, 'Davangere', 'Davangere', 'Davangere Tq', 'Davangere Tq', 'Davangere Dist', 'Davangere Dist', '577001', 'Karnataka', 'India', '', '', _today)]
        cursor.executemany("""INSERT INTO Shops_Donors (Year1,Shop_Donor_ID,Shop_Donor_Name,Place_ID,Place,Place_Kan,Taluk,Taluk_Kan,Dist,Dist_Kan,Pin,State,Country,Mobile,Remarks,DateStamp) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""" if db_type=="postgresql" else """INSERT INTO Shops_Donors (Year1,Shop_Donor_ID,Shop_Donor_Name,Place_ID,Place,Place_Kan,Taluk,Taluk_Kan,Dist,Dist_Kan,Pin,State,Country,Mobile,Remarks,DateStamp) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", _fallback)
        print("Using fallback donor seed.")

    # Dynamic CSV & Excel scanner for grocery items
    import csv
    import os
    base_dir = os.path.dirname(os.path.abspath(__file__))
    file_loaded = False
    files = os.listdir(base_dir)
    
    # Check for Excel files first, then CSV
    xlsx_files = [f for f in files if f.endswith('.xlsx') and not f.startswith('~$')]
    csv_files = [f for f in files if f.endswith('.csv')]
    
    # 1. Try Excel files
    for f_name in xlsx_files:
        xlsx_path = os.path.join(base_dir, f_name)
        try:
            import openpyxl
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
            sheet_name = 'Grocery_Items' if 'Grocery_Items' in wb.sheetnames else wb.sheetnames[0]
            print(f"Detected Excel dataset at '{f_name}' in sheet '{sheet_name}'. Parsing and importing...")
            sheet = wb[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            
            headers = [str(cell).strip() for cell in rows[0] if cell is not None]
            h_map = {h.lower().replace(" ", "_"): idx for idx, h in enumerate(headers)}
            
            # Verify if it looks like a grocery items sheet (has code column)
            if any(h in h_map for h in ['grocery_code', 'code', 'grocerycode']):
                items = []
                for row in rows[1:]:
                    if not row or all(cell is None for cell in row):
                        continue
                    
                    # Code
                    code_idx = h_map.get('grocery_code') or h_map.get('code') or h_map.get('grocerycode')
                    try:
                        code = int(float(row[code_idx]))
                    except:
                        continue
                    
                    # Kannada name
                    name_kan = ""
                    name_idx = h_map.get('grocery_items') or h_map.get('grocery_items_kan') or h_map.get('item_name') or h_map.get('item_name_kannada')
                    if name_idx is not None:
                        name_kan = str(row[name_idx] or '')
                    
                    # English name
                    name_eng = ""
                    name_eng_idx = h_map.get('grocery_items_eng') or h_map.get('item_name_english') or h_map.get('english_name')
                    if name_eng_idx is not None:
                        name_eng = str(row[name_eng_idx] or '')
                        
                    # Category
                    category = "ಧಾನ್ಯಗಳು"
                    cat_idx = h_map.get('grocery_category') or h_map.get('category')
                    if cat_idx is not None:
                        category = str(row[cat_idx] or 'ಧಾನ್ಯಗಳು')
                        
                    # Cat Code
                    cat_code = "C01"
                    cat_code_idx = h_map.get('category_code')
                    if cat_code_idx is not None:
                        cat_code = str(row[cat_code_idx] or 'C01')
                        
                    # Unit
                    unit = "ಕೆಜಿ (KG)"
                    unit_idx = h_map.get('qtl_kg_ltr') or h_map.get('unit')
                    if unit_idx is not None:
                        unit = str(row[unit_idx] or 'ಕೆಜಿ (KG)')
                        
                    # Rate
                    rate = 0.0
                    rate_idx = h_map.get('std_rate') or h_map.get('rate')
                    if rate_idx is not None:
                        try: rate = float(row[rate_idx] or 0.0)
                        except: pass
                        
                    # Helper to get values
                    def get_val(keys, default=0.0):
                        for k in keys:
                            k_mod = k.lower().replace(" ", "_")
                            if k_mod in h_map:
                                val = row[h_map[k_mod]]
                                try: return float(val or default)
                                except: pass
                        return default
                        
                    shraddanjali = get_val(['shraddanjali', 'shraddanjali_qty', 'shraddhajali', 'store_-_shraddhajali'])
                    hunnime = get_val(['hunnime', 'hunnime_qty', 'stores_-_a'])
                    boys = get_val(['boys_hostel', 'boys_hostel_qty', 'boys', 'stores_-_boys_hostel', 'stores_-_sirigere_bhs'])
                    girls = get_val(['girls_hostel', 'girls_hostel_qty', 'girls', 'stores_-_girls_hostel', 'stores_-_sirigere_ghs'])
                    math_val = get_val(['math', 'math_qty', 'stores_-_math'])
                    shantivan = get_val(['shantivana', 'shantivan_qty_a', 'shantivan', 'stores_-_shantivana_bidara', 'stores_-_shantivana_gurukula'])
                    ao_office = get_val(['ao_office', 'ao_office_qty', 'office', 'stores_-_ao_office'])
                    
                    tot_qty = get_val(['tot_quantity', 'total_quantity']) or (shraddanjali + hunnime + boys + girls + math_val + shantivan + ao_office)
                    open_stock = get_val(['opening_stock']) or tot_qty
                    close_stock = get_val(['closing_stock']) or tot_qty
                    tot_stock = get_val(['tot_stock']) or tot_qty
                    tot_issue = get_val(['tot_issue']) or 0.0
                    stock_amt = get_val(['stock_amt']) or (tot_qty * rate)
                    budget = get_val(['budget', 'total_budget']) or (tot_qty * rate * 1.5)
                    remarks_idx = h_map.get('remarks')
                    remarks = str(row[remarks_idx] or '') if remarks_idx is not None else 'Imported from Excel'
                    
                    items.append((
                        code, name_kan, name_eng, category, cat_code, rate, unit,
                        shraddanjali, hunnime, boys, girls, math_val, shantivan, ao_office,
                        shraddanjali, hunnime, boys, girls, math_val, shantivan, ao_office,
                        tot_qty, open_stock, close_stock, tot_stock, tot_issue, stock_amt, budget, remarks, datetime.now().strftime('%Y-%m-%d')
                    ))
                
                if items:
                    cursor.execute("DELETE FROM Grocery_Items;")
                    cursor.executemany("""
                        INSERT INTO Grocery_Items
                        (Grocery_Code, Grocery_Items_Kan, Grocery_Items_Eng, Grocery_Category, Category_Code, Std_Rate, Qtl_Kg_Ltr,
                         Shraddanjali_Qty, Hunnime_Qty, Boys_Hostel_Qty, Girls_Hostel_Qty, Math_Qty, Shantivan_Qty_a, AO_Office_Qty,
                         Shraddanjali_Budget, Hunnime_Budget, Boys_Hostel_Budget, Girls_Hostel_Budget, Math_Budget, Shantivan_Budget_a, AO_Office_Budget,
                         Tot_Quantity, Opening_Stock, Closing_Stock, Tot_Stock, Tot_Issue, Stock_Amt, Total_Budget, Remarks, DateStamp)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """ if db_type == "postgresql" else """
                        INSERT INTO Grocery_Items
                        (Grocery_Code, Grocery_Items_Kan, Grocery_Items_Eng, Grocery_Category, Category_Code, Std_Rate, Qtl_Kg_Ltr,
                         Shraddanjali_Qty, Hunnime_Qty, Boys_Hostel_Qty, Girls_Hostel_Qty, Math_Qty, Shantivan_Qty_a, AO_Office_Qty,
                         Shraddanjali_Budget, Hunnime_Budget, Boys_Hostel_Budget, Girls_Hostel_Budget, Math_Budget, Shantivan_Budget_a, AO_Office_Budget,
                         Tot_Quantity, Opening_Stock, Closing_Stock, Tot_Stock, Tot_Issue, Stock_Amt, Total_Budget, Remarks, DateStamp)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """, items)
                    print(f"Successfully loaded {len(items)} items from Excel sheet: {sheet_name}")
                    file_loaded = True
                    break
        except Exception as e:
            print(f"Failed to parse Excel file '{f_name}': {e}")
            
    # 2. Try CSV files if no Excel loaded
    if not file_loaded:
        for f_name in csv_files:
            csv_path = os.path.join(base_dir, f_name)
            try:
                with open(csv_path, mode='r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    headers = reader.fieldnames
                    if headers and any(h.lower().replace(" ", "_") in ['grocery_code', 'code', 'grocerycode'] for h in headers):
                        print(f"Detected CSV dataset at '{f_name}'. Parsing and importing...")
                        items = []
                        for row in reader:
                            code = 0
                            for h in headers:
                                if h.lower().replace(" ", "_") in ['grocery_code', 'code', 'grocerycode']:
                                    try: code = int(float(row[h]))
                                    except: pass
                            if not code:
                                continue
                            
                            # Kannada name
                            name_kan = ""
                            for h in headers:
                                if h.lower().replace(" ", "_") in ['grocery_items_kan', 'grocery_items_kannada', 'item_name_kannada', 'kannada_name', 'items_kan', 'grocery_items']:
                                    name_kan = row[h]
                            if not name_kan:
                                for h in headers:
                                    if 'name' in h.lower() or 'item' in h.lower():
                                        name_kan = row[h]
                                        break
                                        
                            # English name
                            name_eng = ""
                            for h in headers:
                                if h.lower().replace(" ", "_") in ['grocery_items_eng', 'grocery_items_english', 'item_name_english', 'english_name', 'items_eng']:
                                    name_eng = row[h]
                                    
                            # Category
                            category = "ಧಾನ್ಯಗಳು"
                            for h in headers:
                                if 'category' in h.lower():
                                    category = row[h]
                                    
                            # Unit
                            unit = "ಕೆಜಿ (KG)"
                            for h in headers:
                                if h.lower().replace(" ", "_") in ['qtl_kg_ltr', 'unit', 'unit_type']:
                                    unit = row[h]
                                    
                            # Rate
                            rate = 0.0
                            for h in headers:
                                if h.lower().replace(" ", "_") in ['std_rate', 'rate', 'standard_rate']:
                                    try: rate = float(row[h])
                                    except: pass
                                    
                            # Quantities helper
                            def get_val_csv(keys, default=0.0):
                                for k in keys:
                                    for h in headers:
                                        if h.lower().replace(" ", "_") == k.lower().replace(" ", "_"):
                                            try: return float(row[h] or default)
                                            except: pass
                                return default
                                
                            shraddanjali = get_val_csv(['shraddanjali', 'shraddanjali_qty', 'shraddhajali', 'store_-_shraddhajali'])
                            hunnime = get_val_csv(['hunnime', 'hunnime_qty', 'stores_-_a'])
                            boys = get_val_csv(['boys_hostel', 'boys_hostel_qty', 'boys', 'stores_-_boys_hostel', 'stores_-_sirigere_bhs'])
                            girls = get_val_csv(['girls_hostel', 'girls_hostel_qty', 'girls', 'stores_-_girls_hostel', 'stores_-_sirigere_ghs'])
                            math_val = get_val_csv(['math', 'math_qty', 'stores_-_math'])
                            shantivan = get_val_csv(['shantivana', 'shantivan_qty_a', 'shantivan', 'stores_-_shantivana_bidara', 'stores_-_shantivana_gurukula'])
                            ao_office = get_val_csv(['ao_office', 'ao_office_qty', 'office', 'stores_-_ao_office'])
                            
                            tot_qty = get_val_csv(['tot_quantity', 'total_quantity']) or (shraddanjali + hunnime + boys + girls + math_val + shantivan + ao_office)
                            open_stock = get_val_csv(['opening_stock']) or tot_qty
                            close_stock = get_val_csv(['closing_stock']) or tot_qty
                            tot_stock = get_val_csv(['tot_stock']) or tot_qty
                            tot_issue = get_val_csv(['tot_issue']) or 0.0
                            stock_amt = get_val_csv(['stock_amt']) or (tot_qty * rate)
                            budget = get_val_csv(['budget', 'total_budget']) or (tot_qty * rate * 1.5)
                            remarks = row.get('remarks') or 'Imported from CSV'
                            
                            items.append((
                                code, name_kan, name_eng, category, 'C01', rate, unit,
                                shraddanjali, hunnime, boys, girls, math_val, shantivan, ao_office,
                                shraddanjali, hunnime, boys, girls, math_val, shantivan, ao_office,
                                tot_qty, open_stock, close_stock, tot_stock, tot_issue, stock_amt, budget, remarks, datetime.now().strftime('%Y-%m-%d')
                            ))
                            
                        if items:
                            cursor.execute("DELETE FROM Grocery_Items;")
                            cursor.executemany("""
                                INSERT INTO Grocery_Items
                                (Grocery_Code, Grocery_Items_Kan, Grocery_Items_Eng, Grocery_Category, Category_Code, Std_Rate, Qtl_Kg_Ltr,
                                 Shraddanjali_Qty, Hunnime_Qty, Boys_Hostel_Qty, Girls_Hostel_Qty, Math_Qty, Shantivan_Qty_a, AO_Office_Qty,
                                 Shraddanjali_Budget, Hunnime_Budget, Boys_Hostel_Budget, Girls_Hostel_Budget, Math_Budget, Shantivan_Budget_a, AO_Office_Budget,
                                 Tot_Quantity, Opening_Stock, Closing_Stock, Tot_Stock, Tot_Issue, Stock_Amt, Total_Budget, Remarks, DateStamp)
                                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                            """ if db_type == "postgresql" else """
                                INSERT INTO Grocery_Items
                                (Grocery_Code, Grocery_Items_Kan, Grocery_Items_Eng, Grocery_Category, Category_Code, Std_Rate, Qtl_Kg_Ltr,
                                 Shraddanjali_Qty, Hunnime_Qty, Boys_Hostel_Qty, Girls_Hostel_Qty, Math_Qty, Shantivan_Qty_a, AO_Office_Qty,
                                 Shraddanjali_Budget, Hunnime_Budget, Boys_Hostel_Budget, Girls_Hostel_Budget, Math_Budget, Shantivan_Budget_a, AO_Office_Budget,
                                 Tot_Quantity, Opening_Stock, Closing_Stock, Tot_Stock, Tot_Issue, Stock_Amt, Total_Budget, Remarks, DateStamp)
                                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                            """, items)
                            print(f"Successfully loaded {len(items)} items from CSV file: {f_name}")
                            file_loaded = True
                            break
            except Exception as e:
                print(f"Error parsing CSV '{f_name}': {e}")
                
    if not file_loaded:
        print("No CSV or Excel file found or parsing failed. Seeding default dummy items instead.")
        # Seeding default items list
        items = [
            (101, 'ಅಕ್ಕಿ (Rice)', 'Rice', 'ಧಾನ್ಯಗಳು', 'C01', 45.0, 'ಕೆಜಿ (KG)', 200.0, 100.0, 1500.0, 1200.0, 500.0, 300.0, 100.0, 3900.0, 2000.0, 3900.0, 5000.0, 1100.0, 49500.0, 100000.0, 'ದಿನನಿತ್ಯದ ಅಡುಗೆಗೆ', '2026-07-06'),
            (102, 'ಗೋಧಿ ಹಿಟ್ಟು (Wheat Flour)', 'Wheat Flour', 'ಹಿಟ್ಟು', 'C02', 38.0, 'ಕೆಜಿ (KG)', 50.0, 20.0, 4.0, 250.0, 100.0, 150.0, 50.0, 624.0, 100.0, 624.0, 1000.0, 376.0, 23712.0, 50000.0, 'Low Stock Alert triggered for Boys Hostel (4.0 KG)', '2026-07-06'),
            (103, 'ತೊಗರಿ ಬೇಳε (Toor Dal)', 'Toor Dal', 'ಬೇಳೆಕಾಳು', 'C03', 120.0, 'ಕೆಜಿ (KG)', 10.0, 50.0, 120.0, 90.0, 40.0, 5.0, 10.0, 325.0, 500.0, 325.0, 1000.0, 675.0, 39000.0, 80000.0, 'Low Stock Alert triggered for Shantivana (5.0 KG)', '2026-07-06'),
            (104, 'ಸಕ್ಕರೆ (Sugar)', 'Sugar', 'ಸಿಹಿ ವಸ್ತು', 'C04', 40.0, 'ಕೆಜಿ (KG)', 10.0, 30.0, 80.0, 60.0, 30.0, 40.0, 20.0, 270.0, 200.0, 270.0, 500.0, 230.0, 10800.0, 25000.0, '', '2026-07-06'),
            (105, 'ಅಡುಗೆ ಎಣ್ಣೆ (Cooking Oil)', 'Cooking Oil', 'ಎಣ್ಣೆ', 'C05', 110.0, 'ಲೀಟರ್ (Ltr)', 5.0, 10.0, 90.0, 80.0, 30.0, 20.0, 10.0, 245.0, 100.0, 245.0, 500.0, 255.0, 26950.0, 60000.0, '', '2026-07-06'),
            (106, 'ಪುಡಿ ಉಪ್ಪು (Salt)', 'Salt', 'ಮಸಾಲೆ', 'C06', 15.0, 'ಕೆಜಿ (KG)', 2.0, 5.0, 15.0, 12.0, 10.0, 8.0, 5.0, 57.0, 50.0, 57.0, 100.0, 43.0, 855.0, 5000.0, 'Low Stock Alert triggered for Shraddanjali (2.0 KG)', '2026-07-06'),
            (108, 'ರಾಗಿ (Ragi)', 'Ragi', 'ಧಾನ್ಯಗಳು', 'C01', 35.0, 'ಕೆಜಿ (KG)', 0.0, 0.0, 5.0, 80.0, 15.0, 50.0, 0.0, 150.0, 300.0, 150.0, 500.0, 350.0, 5250.0, 15000.0, 'Low Stock Alert triggered for Boys Hostel (5.0 KG)', '2026-07-06')
        ]
        cursor.executemany("""
            INSERT INTO Grocery_Items
            (Grocery_Code, Grocery_Items_Kan, Grocery_Items_Eng, Grocery_Category, Category_Code, Std_Rate, Qtl_Kg_Ltr,
             Shraddanjali_Qty, Hunnime_Qty, Boys_Hostel_Qty, Girls_Hostel_Qty, Math_Qty, Shantivan_Qty_a, AO_Office_Qty,
             Tot_Quantity, Opening_Stock, Closing_Stock, Tot_Stock, Tot_Issue, Stock_Amt, Total_Budget, Remarks, DateStamp)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """ if db_type == "postgresql" else """
            INSERT INTO Grocery_Items
            (Grocery_Code, Grocery_Items_Kan, Grocery_Items_Eng, Grocery_Category, Category_Code, Std_Rate, Qtl_Kg_Ltr,
             Shraddanjali_Qty, Hunnime_Qty, Boys_Hostel_Qty, Girls_Hostel_Qty, Math_Qty, Shantivan_Qty_a, AO_Office_Qty,
             Tot_Quantity, Opening_Stock, Closing_Stock, Tot_Stock, Tot_Issue, Stock_Amt, Total_Budget, Remarks, DateStamp)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, items)



    # Stock Inward Logs Seed
    inwards = [
        ('2026', '2026-07-01', 101, 2001, None, None, None, None, None, 45.0, 225000.0, 'No', 5000.0, None, 0.0, 0.0, 'ಕೇಂದ್ರ ಗೋದಾಮು', '2026-07-01', 'ಸರ್ಕಾರದಿಂದ ಆಹಾರ ಧಾನ್ಯ ಸರಬರಾಜು', 'Purchase'),
        ('2026', '2026-07-02', 103, 2002, None, None, None, None, None, 120.0, 120000.0, 'No', 1000.0, None, 0.0, 0.0, 'ಕೇಂದ್ರ ಗೋದಾಮು', '2026-07-02', 'ಭಕ್ತರಿಂದ ಕೊಡುಗೆ', 'Donation')
    ]
    cursor.executemany("""
        INSERT INTO Stock_Issue
        (Year1, Date1, Grocery_Code, Shop_Donor_ID, Book_No, Receipt_No, Receipt_Date, Bill_Date, Bill_No, Purchase_Rate, Purchase_Amount, Paid, Stock, Issue_Inst_ID, Issue, Issue_Amount, Received_By, DateStamp, Remarks, Purchased_Donation)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """ if db_type == "postgresql" else """
        INSERT INTO Stock_Issue
        (Year1, Date1, Grocery_Code, Shop_Donor_ID, Book_No, Receipt_No, Receipt_Date, Bill_Date, Bill_No, Purchase_Rate, Purchase_Amount, Paid, Stock, Issue_Inst_ID, Issue, Issue_Amount, Received_By, DateStamp, Remarks, Purchased_Donation)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, inwards)

    # Indents Seeding (Pending, Sent, Received)
    indents = [
        # Pending request
        ('2026', '2026-07-05', 101, 1, 150.0, 'IND-1001', 0.0, 'Pending', None, '2026-07-05', 'ವಾರದ ಅಡುಗೆಗೆ ಬೇಕಾಗಿದೆ'),
        # Sent (Issued but not acknowledged yet)
        ('2026', '2026-07-05', 103, 1, 20.0, 'IND-1002', 20.0, 'Sent', '2026-07-05', '2026-07-05', 'ತುರ್ತು ಬೇಳೆ ಅವಶ್ಯಕತೆ'),
        # Received (Fully completed handshake)
        ('2026', '2026-07-04', 105, 1, 30.0, 'IND-1003', 30.0, 'Received', '2026-07-04', '2026-07-04', 'ತಿಂಗಳ ಅಡುಗೆ ಎಣ್ಣೆ ಕೋಟಾ')
    ]
    cursor.executemany("""
        INSERT INTO Indents
        (Year1, Indent_Date, Grocery_Code, Inst_ID, Quantity, Indent_no, Sanctioned_Quantity, Sanctioned, Sanctioned_on, DateStamp, Remarks)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """ if db_type == "postgresql" else """
        INSERT INTO Indents
        (Year1, Indent_Date, Grocery_Code, Inst_ID, Quantity, Indent_no, Sanctioned_Quantity, Sanctioned, Sanctioned_on, DateStamp, Remarks)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)
    """, indents)

    # Insert corresponding issues for completed or active transfers
    issues = [
        ('2026', '2026-07-05', 103, None, None, None, None, None, None, 120.0, 0.0, 'No', 0.0, 1, 20.0, 2400.0, 'ಗುಮಾಸ್ತ ರಾಮಚಂದ್ರ', '2026-07-05', 'IND-1002 ಅನುಮೋದನೆ', 'Issue'),
        ('2026', '2026-07-04', 105, None, None, None, None, None, None, 110.0, 0.0, 'No', 0.0, 1, 30.0, 3300.0, 'ಗುಮಾಸ್ತ ರಾಮಚಂದ್ರ', '2026-07-04', 'IND-1003 ಸ್ವೀಕರಿಸಲಾಗಿದೆ', 'Issue')
    ]
    cursor.executemany("""
        INSERT INTO Stock_Issue
        (Year1, Date1, Grocery_Code, Shop_Donor_ID, Book_No, Receipt_No, Receipt_Date, Bill_Date, Bill_No, Purchase_Rate, Purchase_Amount, Paid, Stock, Issue_Inst_ID, Issue, Issue_Amount, Received_By, DateStamp, Remarks, Purchased_Donation)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """ if db_type == "postgresql" else """
        INSERT INTO Stock_Issue
        (Year1, Date1, Grocery_Code, Shop_Donor_ID, Book_No, Receipt_No, Receipt_Date, Bill_Date, Bill_No, Purchase_Rate, Purchase_Amount, Paid, Stock, Issue_Inst_ID, Issue, Issue_Amount, Received_By, DateStamp, Remarks, Purchased_Donation)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, issues)

    conn.commit()
    conn.close()
    print(f"Database ({db_type}) successfully initialized and seeded with Kannada values.")

def ensure_audit_logs_table():
    """Create the Audit_Logs table if it doesn't exist (idempotent)."""
    import sqlite3 as _sqlite3
    base_dir = os.path.dirname(os.path.abspath(__file__))
    db_path = os.path.join(base_dir, 'database.db')
    conn = _sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS Audit_Logs (
            Rec INTEGER PRIMARY KEY AUTOINCREMENT,
            Timestamp TEXT NOT NULL,
            Username TEXT NOT NULL,
            Module TEXT NOT NULL,
            Action TEXT NOT NULL,
            Target_ID TEXT,
            Old_Value TEXT,
            New_Value TEXT
        )
    """)
    # Also add Bill_No and Bill_Date columns to Stock_Issue if missing
    try:
        cursor.execute("ALTER TABLE Stock_Issue ADD COLUMN Bill_No TEXT")
    except Exception:
        pass
    try:
        cursor.execute("ALTER TABLE Stock_Issue ADD COLUMN Bill_Date TEXT")
    except Exception:
        pass
    conn.commit()
    conn.close()
    print("Audit_Logs table ensured.")



if __name__ == '__main__':
    init_db()
