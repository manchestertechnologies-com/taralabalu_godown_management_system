import os
import urllib.parse
from flask import Flask, render_template, jsonify, request, session, redirect, url_for
from datetime import datetime
import sqlite3
import uuid

try:
    from PIL import Image
except ImportError:
    Image = None

try:
    import psycopg2
    import psycopg2.extras
except ImportError:
    pass

app = Flask(__name__)
app.secret_key = 'taralabalu_secret_key_87924'

UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'uploads', 'bills')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def compress_image(file_path):
    if Image is None:
        print("Pillow is not installed. Skipping image compression.")
        return
    try:
        ext = os.path.splitext(file_path)[1].lower()
        if ext not in ['.jpg', '.jpeg', '.png']:
            return
        
        with Image.open(file_path) as img:
            # Convert RGBA/P to RGB for saving as JPEG or standard RGB PNG
            if img.mode in ("RGBA", "P"):
                img = img.convert("RGB")
            
            # Resize if width > 1600px
            max_width = 1600
            if img.width > max_width:
                ratio = max_width / float(img.width)
                new_height = int(float(img.height) * ratio)
                img = img.resize((max_width, new_height), Image.Resampling.LANCZOS)
            
            # Save compressed
            if ext in ['.jpg', '.jpeg']:
                img.save(file_path, "JPEG", quality=75, optimize=True)
            elif ext == '.png':
                img.save(file_path, "PNG", optimize=True)
    except Exception as e:
        print(f"Error compressing image {file_path}: {e}")

USERS = {
    "head": {"password": "123", "role": "head", "inst_id": None, "name": "Admin / ಸಂಸ್ಥೆಯ ಆಡಳಿತಗಾರರು"},
    "godown": {"password": "123", "role": "godown", "inst_id": None, "name": "Central Godown Clerk / ಗೋದಾಮು ಗುಮಾಸ್ತ"},
    "accounts": {"password": "123", "role": "accounts", "inst_id": None, "name": "Accounts / ಲೆಕ್ಕಾಧಿಕಾರಿ"},
    "boyshostel": {"password": "123", "role": "hostel", "inst_id": 1, "name": "Stores - Boys Hostel"},
    "girlshostel": {"password": "123", "role": "hostel", "inst_id": 2, "name": "Stores - Girls Hostel"},
    "math": {"password": "123", "role": "hostel", "inst_id": 3, "name": "Stores - Math"},
    "shantivanabidara": {"password": "123", "role": "hostel", "inst_id": 4, "name": "Stores - Shantivana Bidara"},
    "shantivanagurukula": {"password": "123", "role": "hostel", "inst_id": 5, "name": "Stores - Shantivana Gurukula"},
    "sirigerebhs": {"password": "123", "role": "hostel", "inst_id": 6, "name": "Stores - Sirigere BHS"},
    "sirigereghs": {"password": "123", "role": "hostel", "inst_id": 7, "name": "Stores - Sirigere GHS"},
    "storesa": {"password": "123", "role": "hostel", "inst_id": 8, "name": "Stores - A"},
    "aooffice": {"password": "123", "role": "hostel", "inst_id": 9, "name": "Stores - AO_Office"},
    "shraddajali": {"password": "123", "role": "hostel", "inst_id": 10, "name": "Store - Shraddhajali"}
}

INST_ID_TO_COLUMN = {
    1: "Boys_Hostel_Qty", 2: "Girls_Hostel_Qty", 3: "Math_Qty",
    4: "Shantivan_Qty_a", 5: "Shantivan_Qty_a", 6: "Boys_Hostel_Qty",
    7: "Girls_Hostel_Qty", 8: "Hunnime_Qty", 9: "AO_Office_Qty", 10: "Shraddanjali_Qty"
}

INST_ID_TO_BUDGET_COLUMN = {
    1: "Boys_Hostel_Budget", 2: "Girls_Hostel_Budget", 3: "Math_Budget",
    4: "Shantivan_Budget_a", 5: "Shantivan_Budget_a", 6: "Boys_Hostel_Budget",
    7: "Girls_Hostel_Budget", 8: "Hunnime_Budget", 9: "AO_Office_Budget", 10: "Shraddanjali_Budget"
}

def get_db_connection():
    if 'DATABASE_URL' in os.environ:
        conn_str = os.environ['DATABASE_URL']
        conn_str = conn_str.replace(":[Bery8792480218]@", ":Bery8792480218@")
        try:
            result = urllib.parse.urlparse(conn_str)
            conn = psycopg2.connect(
                database=result.path[1:], user=result.username, password=result.password,
                host=result.hostname, port=result.port, connect_timeout=3)
            return conn, "postgresql"
        except Exception:
            pass
    base_dir = os.path.dirname(os.path.abspath(__file__))
    conn = sqlite3.connect(os.path.join(base_dir, 'database.db'))
    return conn, "sqlite"

def db_query(query, args=(), fetch=True):
    conn, db_type = get_db_connection()
    if db_type == "sqlite":
        query = query.replace("%s", "?")
    cursor = conn.cursor()
    try:
        # Use RealDictCursor if postgresql
        if db_type == "postgresql":
            cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cursor.execute(query, args)
        if fetch:
            if db_type == "postgresql":
                result = [dict(row) for row in cursor.fetchall()]
            else:
                if cursor.description:
                    columns = [desc[0] for desc in cursor.description]
                    result = [dict(zip(columns, row)) for row in cursor.fetchall()]
                else:
                    result = []
            cursor.close(); conn.close()
            return result
        else:
            conn.commit(); cursor.close(); conn.close()
            return None
    except Exception as e:
        conn.rollback(); cursor.close(); conn.close()
        raise e

def log_audit(module, action, target_id='', old_val='', new_val=''):
    try:
        username = session.get('username', 'system')
        db_query("INSERT INTO Audit_Logs (Timestamp,Username,Module,Action,Target_ID,Old_Value,New_Value) VALUES (%s,%s,%s,%s,%s,%s,%s)",
            (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), username, module, action, str(target_id), str(old_val)[:500], str(new_val)[:500]), fetch=False)
    except Exception:
        pass

def chk(*roles):
    return 'role' in session and session['role'] in roles

def get_fy_prefix(fy_str):
    if fy_str and len(fy_str) >= 4:
        return fy_str[:4]
    return str(datetime.now().year)

# --- PAGES ---

@app.route('/')
def index():
    if 'username' not in session: return redirect(url_for('login_page'))
    return render_template('index.html')

@app.route('/login')
@app.route('/login/<slug>')
def login_page(slug=None):
    if 'username' in session: return redirect(url_for('index'))
    return render_template('login.html', preselected_slug=slug)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login_page'))

# --- AUTH ---

@app.route('/api/login', methods=['POST'])
def api_login():
    data = request.json or {}
    username = data.get('username', '').strip().lower()
    password = data.get('password', '')
    try:
        users = db_query("SELECT * FROM Users WHERE username = %s", (username,))
        if users and users[0]['password'] == password:
            u = users[0]
            session.update({'username': u['username'], 'role': u['role'], 'inst_id': u['inst_id'], 'name': u['name']})
            log_audit('Auth', 'Login', username)
            return jsonify({'success': True})
    except Exception:
        if username in USERS and USERS[username]['password'] == password:
            session.update({'username': username, 'role': USERS[username]['role'],
                            'inst_id': USERS[username]['inst_id'], 'name': USERS[username]['name']})
            return jsonify({'success': True})
    return jsonify({'success': False, 'message': 'Invalid credentials'}), 401

# --- GROCERY ITEMS ---

@app.route('/api/grocery-items', methods=['GET'])
def get_grocery_items():
    return jsonify(db_query("SELECT * FROM Grocery_Items ORDER BY Grocery_Code"))

@app.route('/api/grocery-items', methods=['POST'])
def add_grocery_item():
    if not chk('head'): return jsonify({'error': 'Unauthorized'}), 403
    d = request.json or {}
    code, name_kan = d.get('Grocery_Code'), d.get('Grocery_Items_Kan')
    if not code or not name_kan: return jsonify({'error': 'Code and Kannada name required'}), 400
    try:
        db_query("INSERT INTO Grocery_Items (Grocery_Code,Grocery_Items_Kan,Grocery_Items_Eng,Grocery_Category,Category_Code,Std_Rate,Qtl_Kg_Ltr,Remarks,DateStamp) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (code, name_kan, d.get('Grocery_Items_Eng',''), d.get('Grocery_Category',''), d.get('Category_Code',''), d.get('Std_Rate',0.0), d.get('Qtl_Kg_Ltr','Kg'), d.get('Remarks',''), datetime.now().strftime('%Y-%m-%d')), fetch=False)
        log_audit('Items', 'Add', code, '', name_kan)
        return jsonify({'success': True})
    except Exception as e: return jsonify({'error': str(e)}), 400

@app.route('/api/grocery-items/<int:code>', methods=['PUT'])
def edit_grocery_item(code):
    if not chk('head'): return jsonify({'error': 'Unauthorized'}), 403
    d = request.json or {}
    try:
        db_query("UPDATE Grocery_Items SET Grocery_Items_Kan=%s,Grocery_Items_Eng=%s,Grocery_Category=%s,Category_Code=%s,Std_Rate=%s,Qtl_Kg_Ltr=%s,Remarks=%s WHERE Grocery_Code=%s",
            (d.get('Grocery_Items_Kan'), d.get('Grocery_Items_Eng',''), d.get('Grocery_Category',''), d.get('Category_Code',''), d.get('Std_Rate',0.0), d.get('Qtl_Kg_Ltr','Kg'), d.get('Remarks',''), code), fetch=False)
        log_audit('Items', 'Edit', code, '', d.get('Grocery_Items_Kan'))
        return jsonify({'success': True})
    except Exception as e: return jsonify({'error': str(e)}), 500

@app.route('/api/grocery-items/<int:code>', methods=['DELETE'])
def delete_grocery_item(code):
    if not chk('head'): return jsonify({'error': 'Unauthorized'}), 403
    try:
        db_query("DELETE FROM Grocery_Items WHERE Grocery_Code = %s", (code,), fetch=False)
        log_audit('Items', 'Delete', code)
        return jsonify({'success': True})
    except Exception as e: return jsonify({'error': str(e)}), 500

# --- INSTITUTIONS ---

@app.route('/api/institutions', methods=['GET'])
def get_institutions():
    return jsonify(db_query("SELECT * FROM Institutions ORDER BY Inst_ID"))

@app.route('/api/institutions', methods=['POST'])
def add_institution():
    if not chk('head'): return jsonify({'error': 'Unauthorized'}), 403
    name = (request.json or {}).get('Institution', '').strip()
    if not name: return jsonify({'error': 'Name required'}), 400
    try:
        db_query("INSERT INTO Institutions (Institution) VALUES (%s)", (name,), fetch=False)
        log_audit('Institutions', 'Add', '', '', name)
        return jsonify({'success': True})
    except Exception as e: return jsonify({'error': str(e)}), 500

# --- DONORS ---

@app.route('/api/donors', methods=['GET'])
def get_donors():
    return jsonify(db_query("SELECT * FROM Shops_Donors ORDER BY Shop_Donor_Name"))

@app.route('/api/donors', methods=['POST'])
def add_donor():
    if not chk('head'): return jsonify({'error':'Unauthorized'}), 403
    d = request.json or {}
    if not d.get('Shop_Donor_ID') or not d.get('Shop_Donor_Name'): return jsonify({'error':'ID and name required'}), 400
    try:
        db_query("INSERT INTO Shops_Donors (Year1,Shop_Donor_ID,Shop_Donor_Name,Place,Mobile,Remarks,DateStamp) VALUES (%s,%s,%s,%s,%s,%s,%s)",
            (datetime.now().year, d['Shop_Donor_ID'], d['Shop_Donor_Name'], d.get('Place',''), d.get('Mobile',''), d.get('Remarks',''), datetime.now().strftime('%Y-%m-%d')), fetch=False)
        log_audit('Donors', 'Add', d['Shop_Donor_ID'], '', d['Shop_Donor_Name'])
        return jsonify({'success': True})
    except Exception as e: return jsonify({'error': str(e)}), 400

@app.route('/api/donors/<donor_id>', methods=['PUT'])
def edit_donor(donor_id):
    if not chk('head'): return jsonify({'error':'Unauthorized'}), 403
    d = request.json or {}
    try:
        db_query("UPDATE Shops_Donors SET Shop_Donor_Name=%s,Place=%s,Mobile=%s,Remarks=%s WHERE Shop_Donor_ID=%s",
            (d.get('Shop_Donor_Name'), d.get('Place',''), d.get('Mobile',''), d.get('Remarks',''), donor_id), fetch=False)
        log_audit('Donors', 'Edit', donor_id, '', d.get('Shop_Donor_Name'))
        return jsonify({'success': True})
    except Exception as e: return jsonify({'error': str(e)}), 500

@app.route('/api/donors/<donor_id>', methods=['DELETE'])
def delete_donor(donor_id):
    if not chk('head'): return jsonify({'error':'Unauthorized'}), 403
    try:
        db_query("DELETE FROM Shops_Donors WHERE Shop_Donor_ID = %s", (donor_id,), fetch=False)
        log_audit('Donors', 'Delete', donor_id)
        return jsonify({'success': True})
    except Exception as e: return jsonify({'error': str(e)}), 500

# --- STOCK INWARD ---

@app.route('/api/godown-stock', methods=['GET'])
def get_godown_stock():
    df = request.args.get('from',''); dt = request.args.get('to',''); dn = request.args.get('donor','')
    fy = request.args.get('year', '')
    inst_id = request.args.get('inst_id') or session.get('inst_id')
    wheres = ["si.Stock > 0"]; args = []
    
    if df: wheres.append("si.Date1 >= %s"); args.append(df)
    if dt: wheres.append("si.Date1 <= %s"); args.append(dt)
    if dn: wheres.append("si.Shop_Donor_ID = %s"); args.append(dn)
    if fy: wheres.append("si.Year1 LIKE %s"); args.append(get_fy_prefix(fy) + "%")
    
    # Store Manager can only view their own store inward entries
    if session.get('role') == 'hostel':
        wheres.append("si.Issue_Inst_ID = %s")
        args.append(inst_id)
    elif inst_id:
        wheres.append("si.Issue_Inst_ID = %s")
        args.append(inst_id)
    else:
        wheres.append("si.Issue_Inst_ID IS NULL")

    logs = db_query(f"""SELECT si.Rec,si.Date1,si.Grocery_Code,gi.Grocery_Items_Kan,gi.Grocery_Items_Eng,
        gi.Qtl_Kg_Ltr,si.Stock,si.Purchase_Rate,(si.Stock*si.Purchase_Rate) AS Amount,
        sd.Shop_Donor_Name,si.Remarks,si.Purchased_Donation,si.Bill_No,si.File_Path
        FROM Stock_Issue si JOIN Grocery_Items gi ON si.Grocery_Code=gi.Grocery_Code
        LEFT JOIN Shops_Donors sd ON si.Shop_Donor_ID=sd.Shop_Donor_ID
        WHERE {' AND '.join(wheres)} ORDER BY si.Rec DESC""", tuple(args))
    return jsonify(logs)

@app.route('/api/godown-stock', methods=['POST'])
def add_godown_stock():
    if not chk('godown', 'hostel', 'head'): return jsonify({'error':'Unauthorized'}), 403
    
    # Multipart form or JSON
    if request.content_type and 'multipart/form-data' in request.content_type:
        code = request.form.get('Grocery_Code')
        donor_id = request.form.get('Shop_Donor_ID')
        quantity = request.form.get('Quantity')
        rate = float(request.form.get('Purchase_Rate') or 0.0)
        remarks = request.form.get('Remarks') or ''
        purch_don = request.form.get('Purchased_Donation') or 'Purchase'
        date_str = request.form.get('Date') or datetime.now().strftime('%Y-%m-%d')
        bill_no = request.form.get('Bill_No') or ''
        bill_date = request.form.get('Bill_Date') or ''
        bill_amount = float(request.form.get('Bill_Amount') or 0.0)
        inst_id = request.form.get('Inst_ID') or session.get('inst_id')
        fy = request.form.get('year') or ''
        bill_file = request.files.get('Bill_File')
    else:
        d = request.json or {}
        code = d.get('Grocery_Code')
        donor_id = d.get('Shop_Donor_ID')
        quantity = d.get('Quantity')
        rate = float(d.get('Purchase_Rate') or 0.0)
        remarks = d.get('Remarks') or ''
        purch_don = d.get('Purchased_Donation') or 'Purchase'
        date_str = d.get('Date') or datetime.now().strftime('%Y-%m-%d')
        bill_no = d.get('Bill_No') or ''
        bill_date = d.get('Bill_Date') or ''
        bill_amount = float(d.get('Bill_Amount') or 0.0)
        inst_id = d.get('Inst_ID') or session.get('inst_id')
        fy = d.get('year') or ''
        bill_file = None

    if not code or not quantity: return jsonify({'error':'Fill all required fields'}), 400
    try:
        qty = float(quantity)
        if qty <= 0: raise ValueError
    except ValueError: return jsonify({'error':'Quantity must be positive'}), 400

    active_year = get_fy_prefix(fy)

    # Handle file upload
    file_path = None
    if bill_file and bill_file.filename:
        ext = os.path.splitext(bill_file.filename)[1].lower()
        if ext not in ['.jpg', '.png', '.jpeg', '.pdf']:
            return jsonify({'error': 'Invalid file format. Supported formats: JPG, PNG, PDF'}), 400
        filename = f"{uuid.uuid4().hex}{ext}"
        bill_file.save(os.path.join(UPLOAD_FOLDER, filename))
        file_path = f"/static/uploads/bills/{filename}"
        compress_image(os.path.join(UPLOAD_FOLDER, filename))

    try:
        # Insert into Stock_Issue
        db_query("""
            INSERT INTO Stock_Issue (Year1, Date1, Grocery_Code, Shop_Donor_ID, Purchase_Rate, Purchase_Amount, Stock, Issue, Remarks, Purchased_Donation, Bill_No, Bill_Date, File_Path, DateStamp, Issue_Inst_ID)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (active_year, date_str, code, donor_id, rate, qty*rate, qty, 0.0, remarks, purch_don, bill_no, bill_date, file_path, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), inst_id), fetch=False)

        # Update Stock levels
        if inst_id:
            col = INST_ID_TO_COLUMN.get(int(inst_id))
            if col:
                db_query(f"UPDATE Grocery_Items SET {col} = COALESCE({col}, 0) + %s WHERE Grocery_Code = %s", (qty, code), fetch=False)
        else:
            db_query("UPDATE Grocery_Items SET Tot_Stock = Tot_Stock + %s WHERE Grocery_Code = %s", (qty, code), fetch=False)

        # Create matching bill if bill number is provided
        if bill_no:
            db_query("""
                INSERT INTO Bills (Year1, Shop_Donor_ID, Bill_Date, Bill_No, Bill_Amount, Remarks, DateAdded, File_Path, Inst_ID)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (active_year, donor_id, bill_date or date_str, bill_no, bill_amount or (qty*rate), remarks, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), file_path, inst_id), fetch=False)

        log_audit('Inward', 'Add', code, '', f'{qty}')
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# --- STOCK OUTWARD ---

@app.route('/api/outwards', methods=['GET'])
def get_outwards():
    df = request.args.get('from',''); dt = request.args.get('to','')
    fy = request.args.get('year', '')
    inst_id = request.args.get('inst_id') or session.get('inst_id')
    wheres = ["si.Issue > 0", "si.Purchased_Donation != 'Consumption'"]; args = []
    
    if df: wheres.append("si.Date1 >= %s"); args.append(df)
    if dt: wheres.append("si.Date1 <= %s"); args.append(dt)
    if fy: wheres.append("si.Year1 LIKE %s"); args.append(get_fy_prefix(fy) + "%")
    
    # Store Manager can only view their own store outward entries
    if session.get('role') == 'hostel':
        wheres.append("si.Issue_Inst_ID = %s")
        args.append(inst_id)
    elif inst_id:
        wheres.append("si.Issue_Inst_ID = %s")
        args.append(inst_id)
    else:
        wheres.append("si.Issue_Inst_ID IS NULL")

    logs = db_query(f"""SELECT si.Rec,si.Date1,si.Grocery_Code,gi.Grocery_Items_Kan,gi.Grocery_Items_Eng,
        gi.Qtl_Kg_Ltr,si.Issue,si.Purchase_Rate,(si.Issue*si.Purchase_Rate) AS Amount,
        si.Remarks,si.Purchased_Donation
        FROM Stock_Issue si JOIN Grocery_Items gi ON si.Grocery_Code=gi.Grocery_Code
        WHERE {' AND '.join(wheres)} ORDER BY si.Rec DESC""", tuple(args))
    return jsonify(logs)

@app.route('/api/outwards', methods=['POST'])
def add_outward():
    if not chk('godown', 'hostel', 'head'): return jsonify({'error':'Unauthorized'}), 403
    d = request.json or {}
    code = d.get('Grocery_Code')
    quantity = d.get('Quantity')
    remarks = d.get('Remarks') or 'Outward'
    date_str = d.get('Date') or datetime.now().strftime('%Y-%m-%d')
    inst_id = d.get('Inst_ID') or session.get('inst_id')
    target_inst_id = d.get('Target_Inst_ID') # for store-to-store transfer
    fy = d.get('year') or ''
    
    if not code or not quantity: return jsonify({'error':'Item and Quantity are required'}), 400
    try:
        qty = float(quantity)
        if qty <= 0: raise ValueError
    except ValueError: return jsonify({'error':'Quantity must be positive'}), 400

    active_year = get_fy_prefix(fy)

    try:
        # Check current stock balance before issuing
        if inst_id:
            col = INST_ID_TO_COLUMN.get(int(inst_id))
            if not col: return jsonify({'error':'Invalid hostel'}), 400
            row = db_query(f"SELECT {col}, Std_Rate FROM Grocery_Items WHERE Grocery_Code=%s", (code,))
            avail = row[0][col] if row and row[0][col] is not None else 0.0
            rate = row[0]['Std_Rate'] or 0.0
        else:
            row = db_query("SELECT Tot_Stock, Std_Rate FROM Grocery_Items WHERE Grocery_Code=%s", (code,))
            avail = row[0]['Tot_Stock'] if row and row[0]['Tot_Stock'] is not None else 0.0
            rate = row[0]['Std_Rate'] or 0.0

        if avail < qty:
            return jsonify({'error': f'Insufficient stock. Available: {avail}'}), 400

        # Insert into Stock_Issue
        db_query("""
            INSERT INTO Stock_Issue (Year1, Date1, Grocery_Code, Issue_Inst_ID, Issue, Issue_Amount, Purchase_Rate, Remarks, Purchased_Donation, DateStamp)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (active_year, date_str, code, inst_id, qty, qty*rate, rate, remarks, 'Transfer' if target_inst_id else 'Outward', datetime.now().strftime('%Y-%m-%d %H:%M:%S')), fetch=False)

        # Update Grocery_Items stock levels (subtract from source)
        if inst_id:
            col = INST_ID_TO_COLUMN.get(int(inst_id))
            db_query(f"UPDATE Grocery_Items SET {col} = {col} - %s WHERE Grocery_Code = %s", (qty, code), fetch=False)
        else:
            db_query("UPDATE Grocery_Items SET Tot_Stock = Tot_Stock - %s, Tot_Issue = Tot_Issue + %s WHERE Grocery_Code = %s", (qty, qty, code), fetch=False)

        # If it is a store-to-store transfer, add to target store
        if target_inst_id:
            target_col = INST_ID_TO_COLUMN.get(int(target_inst_id))
            if target_col:
                db_query(f"UPDATE Grocery_Items SET {target_col} = COALESCE({target_col}, 0) + %s WHERE Grocery_Code = %s", (qty, code), fetch=False)
                # Also record receiving entry
                db_query("""
                    INSERT INTO Stock_Issue (Year1, Date1, Grocery_Code, Issue_Inst_ID, Stock, Purchase_Rate, Purchase_Amount, Remarks, Purchased_Donation, DateStamp)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (active_year, date_str, code, target_inst_id, qty, rate, qty*rate, f"Transfer from inst {inst_id}", 'Transfer_Receipt', datetime.now().strftime('%Y-%m-%d %H:%M:%S')), fetch=False)

        log_audit('Outward', 'Add', code, '', f'{qty}')
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# --- INDENTS ---

@app.route('/api/indents', methods=['GET'])
def get_indents():
    hid = request.args.get('hostel_id'); sf = request.args.get('status','')
    df = request.args.get('from',''); dt = request.args.get('to','')
    fy = request.args.get('year', '')
    wheres = []; args = []
    
    # Store Manager can only view their own store data
    if session.get('role') == 'hostel':
        wheres.append("ind.Inst_ID=%s")
        args.append(session.get('inst_id'))
    elif hid:
        wheres.append("ind.Inst_ID=%s")
        args.append(hid)
        
    if sf: wheres.append("ind.Sanctioned=%s"); args.append(sf)
    if df: wheres.append("ind.Indent_Date>=%s"); args.append(df)
    if dt: wheres.append("ind.Indent_Date<=%s"); args.append(dt)
    if fy: wheres.append("ind.Year1 LIKE %s"); args.append(get_fy_prefix(fy) + "%")
    
    where_sql = ("WHERE " + " AND ".join(wheres)) if wheres else ""
    logs = db_query(f"""SELECT ind.Rec,ind.Indent_Date,ind.Grocery_Code,
        gi.Grocery_Items_Kan,gi.Grocery_Items_Eng,gi.Qtl_Kg_Ltr,
        ind.Quantity,ind.Sanctioned_Quantity,ind.Sanctioned,ind.Sanctioned_on,
        ind.Indent_no,ind.Remarks,inst.Institution,ind.Inst_ID
        FROM Indents ind JOIN Grocery_Items gi ON ind.Grocery_Code=gi.Grocery_Code
        JOIN Institutions inst ON ind.Inst_ID=inst.Inst_ID
        {where_sql} ORDER BY ind.Rec DESC""", tuple(args))
    return jsonify(logs)

@app.route('/api/indents', methods=['POST'])
def create_indent():
    if not chk('head','hostel','godown'): return jsonify({'error':'Unauthorized'}), 403
    d = request.json or {}
    inst_id = d.get('Inst_ID') or session.get('inst_id')
    code, quantity = d.get('Grocery_Code'), d.get('Quantity')
    fy = d.get('year') or ''
    if not inst_id or not code or not quantity: return jsonify({'error':'Fill all required fields'}), 400
    try:
        qty = float(quantity)
        if qty <= 0: raise ValueError
    except ValueError: return jsonify({'error':'Quantity must be positive'}), 400
    active_year = get_fy_prefix(fy)
    indent_no = f"IND-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    
    # Check hostel item budget allocation limit
    budget_col = INST_ID_TO_BUDGET_COLUMN.get(int(inst_id))
    if budget_col:
        item_rows = db_query(f"SELECT {budget_col} AS budget, Grocery_Items_Kan FROM Grocery_Items WHERE Grocery_Code=%s", (code,))
        if item_rows:
            item_info = item_rows[0]
            budget_val = item_info['budget'] or 0.0
            if budget_val > 0.0:
                util_rows = db_query("SELECT SUM(Quantity) AS total_requested FROM Indents WHERE Inst_ID=%s AND Grocery_Code=%s AND Sanctioned IN ('Sent', 'Received', 'Pending') AND Year1=%s", (inst_id, code, active_year))
                util_val = (util_rows[0]['total_requested'] or 0.0) if util_rows else 0.0
                if util_val + qty > budget_val:
                    return jsonify({'error': f"Exceeds item budget limit. Allocated: {budget_val}, Used/Pending: {util_val}, Remaining: {max(0.0, budget_val - util_val)}"}), 400

    try:
        db_query("INSERT INTO Indents (Year1,Indent_Date,Grocery_Code,Inst_ID,Quantity,Indent_no,Sanctioned,Remarks,DateStamp) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (active_year, datetime.now().strftime('%Y-%m-%d'), code, inst_id, qty, indent_no, 'Pending', d.get('Remarks',''), datetime.now().strftime('%Y-%m-%d')), fetch=False)
        log_audit('Indents', 'Create', indent_no, '', f'{qty} of {code}')
        return jsonify({'success': True, 'indent_no': indent_no})
    except Exception as e: return jsonify({'error': str(e)}), 500

@app.route('/api/indents/<int:indent_rec>/approve', methods=['POST'])
def approve_indent(indent_rec):
    if not chk('godown'): return jsonify({'error':'Unauthorized'}), 403
    d = request.json or {}
    rows = db_query("SELECT * FROM Indents WHERE Rec=%s", (indent_rec,))
    if not rows: return jsonify({'error':'Indent not found'}), 404
    indent = rows[0]
    if indent['Sanctioned'] != 'Pending': return jsonify({'error':'Already processed'}), 400
    sq = float(d.get('Sanctioned_Quantity') or indent['Quantity'])
    if sq <= 0: return jsonify({'error':'Invalid quantity'}), 400
    code = indent['Grocery_Code']
    item_rows = db_query("SELECT Tot_Stock,Std_Rate FROM Grocery_Items WHERE Grocery_Code=%s", (code,))
    if not item_rows: return jsonify({'error':'Item not found'}), 404
    item = item_rows[0]
    if (item['Tot_Stock'] or 0) < sq: return jsonify({'error':f'Insufficient stock. Available: {item["Tot_Stock"]}'}), 400
    try:
        db_query("UPDATE Indents SET Sanctioned_Quantity=%s,Sanctioned='Sent',Sanctioned_on=%s WHERE Rec=%s",
            (sq, datetime.now().strftime('%Y-%m-%d'), indent_rec), fetch=False)
        rate = item['Std_Rate'] or 0.0
        db_query("INSERT INTO Stock_Issue (Year1,Date1,Grocery_Code,Issue_Inst_ID,Issue,Issue_Amount,Remarks,Purchased_Donation,DateStamp) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (indent['Year1'], datetime.now().strftime('%Y-%m-%d'), code, indent['Inst_ID'], sq, sq*rate, f"Indent {indent['Indent_no']}", 'Issue', datetime.now().strftime('%Y-%m-%d')), fetch=False)
        db_query("UPDATE Grocery_Items SET Tot_Stock=Tot_Stock-%s,Tot_Issue=Tot_Issue+%s WHERE Grocery_Code=%s", (sq,sq,code), fetch=False)
        log_audit('Indents', 'Approve', indent['Indent_no'], indent['Quantity'], sq)
        return jsonify({'success': True})
    except Exception as e: return jsonify({'error': str(e)}), 500

@app.route('/api/indents/<int:indent_rec>/reject', methods=['POST'])
def reject_indent(indent_rec):
    if not chk('godown'): return jsonify({'error':'Unauthorized'}), 403
    d = request.json or {}
    rows = db_query("SELECT * FROM Indents WHERE Rec=%s", (indent_rec,))
    if not rows: return jsonify({'error':'Indent not found'}), 404
    if rows[0]['Sanctioned'] != 'Pending': return jsonify({'error':'Already processed'}), 400
    reason = d.get('Reason', 'Rejected')
    try:
        db_query("UPDATE Indents SET Sanctioned='Rejected',Sanctioned_on=%s,Remarks=%s WHERE Rec=%s",
            (datetime.now().strftime('%Y-%m-%d'), reason, indent_rec), fetch=False)
        log_audit('Indents', 'Reject', indent_rec, '', reason)
        return jsonify({'success': True})
    except Exception as e: return jsonify({'error': str(e)}), 500

@app.route('/api/indents/<int:indent_rec>/acknowledge', methods=['POST'])
def acknowledge_indent(indent_rec):
    if not chk('head','hostel','godown'): return jsonify({'error':'Unauthorized'}), 403
    rows = db_query("SELECT * FROM Indents WHERE Rec=%s", (indent_rec,))
    if not rows: return jsonify({'error':'Not found'}), 404
    indent = rows[0]
    if indent['Sanctioned'] == 'Received': return jsonify({'error':'Already acknowledged'}), 400
    col = INST_ID_TO_COLUMN.get(int(indent['Inst_ID']))
    if not col: return jsonify({'error':'Unmapped institution'}), 400
    try:
        db_query("UPDATE Indents SET Sanctioned='Received',Sanctioned_on=%s WHERE Rec=%s",
            (datetime.now().strftime('%Y-%m-%d'), indent_rec), fetch=False)
        db_query(f"UPDATE Grocery_Items SET {col}={col}+%s WHERE Grocery_Code=%s",
            (indent['Sanctioned_Quantity'], indent['Grocery_Code']), fetch=False)
        log_audit('Indents', 'Acknowledge', indent['Indent_no'], '', indent['Sanctioned_Quantity'])
        return jsonify({'success': True})
    except Exception as e: return jsonify({'error': str(e)}), 500

# --- HOSTEL STOCK ---

@app.route('/api/hostel-stock/<int:inst_id>', methods=['GET'])
def get_hostel_stock(inst_id):
    # Store manager can only view their own store data
    if session.get('role') == 'hostel' and session.get('inst_id') != inst_id:
        return jsonify({'error':'Unauthorized'}), 403
    col = INST_ID_TO_COLUMN.get(inst_id)
    budget_col = INST_ID_TO_BUDGET_COLUMN.get(inst_id)
    if not col: return jsonify([])
    select_budget = f"{budget_col} AS AllocatedBudget" if budget_col else "0.0 AS AllocatedBudget"
    return jsonify(db_query(f"SELECT Grocery_Code,Grocery_Items_Kan,Grocery_Items_Eng,Grocery_Category,Qtl_Kg_Ltr,{col} AS CurrentBalance,{select_budget},Std_Rate FROM Grocery_Items WHERE {col}>=0 ORDER BY Grocery_Category,Grocery_Code"))

# --- DAILY USAGE ---

@app.route('/api/daily-usage', methods=['POST'])
def add_daily_usage():
    if not chk('head','hostel','godown'): return jsonify({'error':'Unauthorized'}), 403
    d = request.json or {}
    inst_id = d.get('Inst_ID') or session.get('inst_id')
    code, quantity = d.get('Grocery_Code'), d.get('Quantity')
    fy = d.get('year') or ''
    if not inst_id or not code or not quantity: return jsonify({'error':'Fill all required fields'}), 400
    try:
        qty = float(quantity)
        if qty <= 0: raise ValueError
    except ValueError: return jsonify({'error':'Quantity must be positive'}), 400
    col = INST_ID_TO_COLUMN.get(int(inst_id))
    if not col: return jsonify({'error':'Invalid hostel'}), 400
    rows = db_query(f"SELECT {col},Std_Rate FROM Grocery_Items WHERE Grocery_Code=%s", (code,))
    if not rows: return jsonify({'error':'Item not found'}), 404
    item = rows[0]
    if (item[col] or 0) < qty: return jsonify({'error':f'Insufficient. Available: {item[col]}'}), 400
    date_str = d.get('Date') or datetime.now().strftime('%Y-%m-%d')
    active_year = get_fy_prefix(fy)
    try:
        db_query(f"UPDATE Grocery_Items SET {col}={col}-%s WHERE Grocery_Code=%s", (qty,code), fetch=False)
        rate = item['Std_Rate'] or 0.0
        db_query("INSERT INTO Stock_Issue (Year1,Date1,Grocery_Code,Issue_Inst_ID,Issue,Issue_Amount,Remarks,Purchased_Donation,DateStamp) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (active_year, date_str, code, inst_id, qty, qty*rate, d.get('Remarks','Consumption'), 'Consumption', datetime.now().strftime('%Y-%m-%d')), fetch=False)
        log_audit('Consumption', 'Add', code, '', f'{qty} by inst {inst_id}')
        return jsonify({'success': True})
    except Exception as e: return jsonify({'error': str(e)}), 500

@app.route('/api/daily-usage/<int:inst_id>', methods=['GET'])
def get_daily_usage_logs(inst_id):
    # Store manager can only view their own store data
    if session.get('role') == 'hostel' and session.get('inst_id') != inst_id:
        return jsonify({'error':'Unauthorized'}), 403
    df = request.args.get('from',''); dt = request.args.get('to','')
    fy = request.args.get('year', '')
    extra = ""; args = [inst_id]
    if df: extra += " AND si.Date1>=%s"; args.append(df)
    if dt: extra += " AND si.Date1<=%s"; args.append(dt)
    if fy: extra += " AND si.Year1 LIKE %s"; args.append(get_fy_prefix(fy) + "%")
    return jsonify(db_query(f"""SELECT si.Rec,si.Date1,si.Grocery_Code,gi.Grocery_Items_Kan,gi.Grocery_Items_Eng,
        gi.Qtl_Kg_Ltr,si.Issue AS QuantityUsed,si.Issue_Amount,si.Remarks
        FROM Stock_Issue si JOIN Grocery_Items gi ON si.Grocery_Code=gi.Grocery_Code
        WHERE si.Issue_Inst_ID=%s AND si.Purchased_Donation='Consumption' {extra}
        ORDER BY si.Rec DESC""", tuple(args)))

@app.route('/api/daily-usage/<int:rec_id>', methods=['DELETE'])
def delete_daily_usage(rec_id):
    if not chk('head', 'hostel', 'godown'): return jsonify({'error':'Unauthorized'}), 403
    rows = db_query("SELECT * FROM Stock_Issue WHERE Rec=%s AND Purchased_Donation='Consumption'", (rec_id,))
    if not rows: return jsonify({'error':'Log not found'}), 404
    log = rows[0]

    # Verify same day entry
    today = datetime.now().strftime('%Y-%m-%d')
    if log['Date1'] != today:
        return jsonify({'error':'Can only edit or delete entries made on the same day!'}), 400

    inst_id = log['Issue_Inst_ID']
    col = INST_ID_TO_COLUMN.get(int(inst_id))
    if not col: return jsonify({'error':'Invalid hostel'}), 400

    try:
        # Revert the stock level by adding back the consumed quantity
        db_query(f"UPDATE Grocery_Items SET {col} = {col} + %s WHERE Grocery_Code = %s", (log['Issue'], log['Grocery_Code']), fetch=False)
        # Delete consumption log
        db_query("DELETE FROM Stock_Issue WHERE Rec = %s", (rec_id,), fetch=False)
        log_audit('Consumption', 'Delete', log['Grocery_Code'], f"{log['Issue']} reverted", 'Deleted')
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# --- LOW STOCK ALERTS ---

@app.route('/api/low-stock-alerts', methods=['GET'])
def get_low_stock_alerts():
    alerts = []
    role = session.get('role')
    user_inst = session.get('inst_id')
    for inst_id, col in INST_ID_TO_COLUMN.items():
        if role == 'hostel' and user_inst != inst_id:
            continue
        inst_rows = db_query("SELECT Institution FROM Institutions WHERE Inst_ID=%s", (inst_id,))
        if not inst_rows: continue
        inst_name = inst_rows[0]['Institution']
        for item in db_query(f"SELECT Grocery_Code,Grocery_Items_Kan,Qtl_Kg_Ltr,{col} AS CurrentBalance FROM Grocery_Items WHERE {col}<10.0 AND {col}>=0"):
            alerts.append({'Inst_ID':inst_id,'Institution':inst_name,'Grocery_Code':item['Grocery_Code'],'Grocery_Items_Kan':item['Grocery_Items_Kan'],'Qtl_Kg_Ltr':item['Qtl_Kg_Ltr'],'CurrentBalance':item['CurrentBalance']})
    return jsonify(alerts)

# --- VEGETABLES ---

@app.route('/api/vegetables', methods=['GET'])
def get_vegetables():
    iid = request.args.get('inst_id'); df = request.args.get('from',''); dt = request.args.get('to','')
    fy = request.args.get('year', '')
    wheres = []; args = []
    
    # Store manager can only view their own store data
    if session.get('role') == 'hostel':
        wheres.append("v.Inst_ID=%s")
        args.append(session.get('inst_id'))
    elif iid:
        wheres.append("v.Inst_ID=%s")
        args.append(iid)
        
    if df: wheres.append("v.Purchase_On>=%s"); args.append(df)
    if dt: wheres.append("v.Purchase_On<=%s"); args.append(dt)
    if fy: wheres.append("v.Year1 LIKE %s"); args.append(get_fy_prefix(fy) + "%")
    
    where_sql = ("WHERE " + " AND ".join(wheres)) if wheres else ""
    return jsonify(db_query(f"SELECT v.*,i.Institution FROM Vegetable v JOIN Institutions i ON v.Inst_ID=i.Inst_ID {where_sql} ORDER BY v.Rec DESC", tuple(args)))

@app.route('/api/vegetables', methods=['POST'])
def add_vegetable():
    if not chk('head','godown','hostel'): return jsonify({'error':'Unauthorized'}), 403
    d = request.json or {}
    fy = d.get('year') or ''
    active_year = get_fy_prefix(fy)
    try:
        db_query("INSERT INTO Vegetable (Inst_ID,Year1,Purchase_On,V_Code,Quantity,Bill_Date,Bill_No,Rate,Remarks,Issue_Place,Purchased_Donation) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (d.get('Inst_ID') or session.get('inst_id'), active_year, d.get('Purchase_On'), d.get('V_Code'), d.get('Quantity',0.0), d.get('Bill_Date'), d.get('Bill_No'), d.get('Rate',0.0), d.get('Remarks'), d.get('Issue_Place'), d.get('Purchased_Donation','Purchase')), fetch=False)
        log_audit('Vegetables','Add',d.get('V_Code',''),'',d.get('Quantity',''))
        return jsonify({'success': True})
    except Exception as e: return jsonify({'error': str(e)}), 500

# --- BILLS ---

@app.route('/api/bills', methods=['GET'])
def get_bills():
    df = request.args.get('from',''); dt = request.args.get('to',''); dn = request.args.get('donor','')
    fy = request.args.get('year', '')
    inst_id = request.args.get('inst_id') or session.get('inst_id')
    wheres = []; args = []
    
    # Store Manager can only view their own store data
    if session.get('role') == 'hostel':
        wheres.append("b.Inst_ID=%s")
        args.append(inst_id)
    elif inst_id:
        wheres.append("b.Inst_ID=%s")
        args.append(inst_id)
        
    if df: wheres.append("b.Bill_Date>=%s"); args.append(df)
    if dt: wheres.append("b.Bill_Date<=%s"); args.append(dt)
    if dn: wheres.append("b.Shop_Donor_ID=%s"); args.append(dn)
    if fy: wheres.append("b.Year1 LIKE %s"); args.append(get_fy_prefix(fy) + "%")
    
    where_sql = ("WHERE " + " AND ".join(wheres)) if wheres else ""
    return jsonify(db_query(f"SELECT b.*,sd.Shop_Donor_Name,sd.Place,inst.Institution FROM Bills b LEFT JOIN Shops_Donors sd ON b.Shop_Donor_ID=sd.Shop_Donor_ID LEFT JOIN Institutions inst ON b.Inst_ID=inst.Inst_ID {where_sql} ORDER BY b.Rec DESC", tuple(args)))

@app.route('/api/bills', methods=['POST'])
def add_bill():
    if not chk('head','accounts','godown','hostel'): return jsonify({'error':'Unauthorized'}), 403
    
    # Handle multipart form or JSON
    if request.content_type and 'multipart/form-data' in request.content_type:
        donor_id = request.form.get('Shop_Donor_ID')
        bill_date = request.form.get('Bill_Date')
        bill_no = request.form.get('Bill_No')
        amount = float(request.form.get('Bill_Amount') or 0.0)
        paid_by = request.form.get('Paid_By') or ''
        ch_date = request.form.get('Ch_Date') or ''
        ch_no = request.form.get('Ch_No') or ''
        ch_amount = float(request.form.get('Ch_Amount') or 0.0)
        remarks = request.form.get('Remarks') or ''
        inst_id = request.form.get('Inst_ID') or session.get('inst_id')
        fy = request.form.get('year') or ''
        bill_file = request.files.get('Bill_File')
    else:
        d = request.json or {}
        donor_id = d.get('Shop_Donor_ID')
        bill_date = d.get('Bill_Date')
        bill_no = d.get('Bill_No')
        amount = float(d.get('Bill_Amount') or 0.0)
        paid_by = d.get('Paid_By') or ''
        ch_date = d.get('Ch_Date') or ''
        ch_no = d.get('Ch_No') or ''
        ch_amount = float(d.get('Ch_Amount') or 0.0)
        remarks = d.get('Remarks') or ''
        inst_id = d.get('Inst_ID') or session.get('inst_id')
        fy = d.get('year') or ''
        bill_file = None

    if not bill_no or not donor_id: return jsonify({'error':'Bill number and Donor/Vendor required'}), 400
    
    active_year = get_fy_prefix(fy)

    # Handle file upload
    file_path = None
    if bill_file and bill_file.filename:
        ext = os.path.splitext(bill_file.filename)[1].lower()
        if ext not in ['.jpg', '.png', '.jpeg', '.pdf']:
            return jsonify({'error': 'Invalid file format. Supported formats: JPG, PNG, PDF'}), 400
        filename = f"{uuid.uuid4().hex}{ext}"
        bill_file.save(os.path.join(UPLOAD_FOLDER, filename))
        file_path = f"/static/uploads/bills/{filename}"
        compress_image(os.path.join(UPLOAD_FOLDER, filename))

    try:
        db_query("INSERT INTO Bills (Year1,Shop_Donor_ID,Bill_Date,Bill_No,Bill_Amount,Paid_By,Ch_Date,Ch_No,Ch_Amount,Remarks,DateAdded,File_Path,Inst_ID) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (active_year, donor_id, bill_date, bill_no, amount, paid_by, ch_date, ch_no, ch_amount, remarks, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), file_path, inst_id), fetch=False)
        log_audit('Bills','Add',bill_no,'',str(amount))
        return jsonify({'success': True})
    except Exception as e: return jsonify({'error': str(e)}), 500

@app.route('/api/bills/<int:rec>', methods=['DELETE'])
def delete_bill(rec):
    if not chk('head','accounts'): return jsonify({'error':'Unauthorized'}), 403
    try:
        db_query("DELETE FROM Bills WHERE Rec=%s", (rec,), fetch=False)
        log_audit('Bills','Delete',rec)
        return jsonify({'success': True})
    except Exception as e: return jsonify({'error': str(e)}), 500


# --- BUDGET MANAGEMENT ---

@app.route('/api/budgets/store', methods=['GET'])
def get_store_budgets():
    year = request.args.get('year') or str(datetime.now().year)
    year_prefix = get_fy_prefix(year)
    
    # Store Manager can only view their own store budget data
    role = session.get('role')
    user_inst = session.get('inst_id')
    
    # Get all institutions (filter for hostel role)
    if role == 'hostel' and user_inst:
        stores_raw = db_query("SELECT Inst_ID, Institution FROM Institutions WHERE Inst_ID = %s ORDER BY Inst_ID", (user_inst,))
    else:
        stores_raw = db_query("SELECT Inst_ID, Institution FROM Institutions ORDER BY Inst_ID")

    # Compute sanctioned budget PER STORE from Grocery_Items budget columns
    # Each store's sanctioned budget = SUM(store_budget_qty * Std_Rate) across all items
    sanctioned_map = {}
    for iid, budget_col in INST_ID_TO_BUDGET_COLUMN.items():
        rows = db_query(f"SELECT COALESCE(SUM({budget_col} * Std_Rate), 0) AS sanctioned FROM Grocery_Items WHERE {budget_col} > 0")
        sanctioned_map[iid] = rows[0]['sanctioned'] if rows else 0.0

    # Utilized budget calculation from actual issues this year
    grocery_spends = db_query("""
        SELECT Issue_Inst_ID, SUM(Issue_Amount) AS spent
        FROM Stock_Issue
        WHERE Purchased_Donation = 'Issue' AND Year1 LIKE %s AND Issue_Inst_ID IS NOT NULL
        GROUP BY Issue_Inst_ID
    """, (year_prefix + '%',))
    g_spend_map = {row['Issue_Inst_ID']: row['spent'] for row in grocery_spends}
    
    veg_spends = db_query("""
        SELECT Inst_ID, SUM(Quantity * Rate) AS spent
        FROM Vegetable
        WHERE Year1 LIKE %s AND Inst_ID IS NOT NULL
        GROUP BY Inst_ID
    """, (year_prefix + '%',))
    v_spend_map = {row['Inst_ID']: row['spent'] for row in veg_spends}
    
    stores = []
    for s in stores_raw:
        iid = s['Inst_ID']
        sanctioned = sanctioned_map.get(iid, 0.0)
        spent = (g_spend_map.get(iid, 0.0) or 0) + (v_spend_map.get(iid, 0.0) or 0)
        stores.append({
            'Inst_ID': iid,
            'Institution': s['Institution'],
            'Allocated_Amount': sanctioned,       # Sanctioned from item budget × rate
            'Used_Amount': spent,
            'Remaining_Amount': sanctioned - spent
        })
        
    return jsonify(stores)


@app.route('/api/budgets/store-items/<int:inst_id>', methods=['GET'])
def get_store_item_budgets(inst_id):
    """Return item-wise sanctioned budget list for a specific store."""
    if session.get('role') == 'hostel' and session.get('inst_id') != inst_id:
        return jsonify({'error': 'Unauthorized'}), 403
    
    budget_col = INST_ID_TO_BUDGET_COLUMN.get(inst_id)
    qty_col = INST_ID_TO_COLUMN.get(inst_id)
    if not budget_col:
        return jsonify([])
    
    year = request.args.get('year') or str(datetime.now().year)
    year_prefix = get_fy_prefix(year)

    # Item-level sanctioned quantities and amounts
    items = db_query(f"""
        SELECT Grocery_Code, Grocery_Items_Kan, Grocery_Items_Eng, Grocery_Category,
               Qtl_Kg_Ltr, Std_Rate,
               {budget_col} AS Sanctioned_Qty,
               {budget_col} * Std_Rate AS Sanctioned_Amount,
               {qty_col} AS Current_Stock
        FROM Grocery_Items
        WHERE {budget_col} > 0
        ORDER BY Grocery_Category, Grocery_Code
    """)

    # Actual issues per item for this store this year
    issues = db_query("""
        SELECT Grocery_Code, SUM(Issue) AS Issued_Qty, SUM(Issue_Amount) AS Issued_Amount
        FROM Stock_Issue
        WHERE Issue_Inst_ID = %s AND Purchased_Donation = 'Issue' AND Year1 LIKE %s
        GROUP BY Grocery_Code
    """, (inst_id, year_prefix + '%'))
    issue_map = {r['Grocery_Code']: r for r in issues}

    result = []
    for item in items:
        code = item['Grocery_Code']
        iss = issue_map.get(code, {})
        sanctioned_qty = item['Sanctioned_Qty'] or 0
        issued_qty = iss.get('Issued_Qty') or 0
        issued_amt = iss.get('Issued_Amount') or 0
        sanctioned_amt = item['Sanctioned_Amount'] or 0
        result.append({
            'Grocery_Code': code,
            'Grocery_Items_Kan': item['Grocery_Items_Kan'],
            'Grocery_Items_Eng': item['Grocery_Items_Eng'],
            'Grocery_Category': item['Grocery_Category'],
            'Qtl_Kg_Ltr': item['Qtl_Kg_Ltr'],
            'Std_Rate': item['Std_Rate'],
            'Sanctioned_Qty': sanctioned_qty,
            'Sanctioned_Amount': sanctioned_amt,
            'Issued_Qty': issued_qty,
            'Issued_Amount': issued_amt,
            'Remaining_Qty': sanctioned_qty - issued_qty,
            'Remaining_Amount': sanctioned_amt - issued_amt,
            'Current_Stock': item['Current_Stock'] or 0,
            'Pct_Used': round((issued_qty / sanctioned_qty * 100), 1) if sanctioned_qty > 0 else 0
        })
    return jsonify(result)


@app.route('/api/budgets/store', methods=['POST'])
def set_store_budget():
    if not chk('head'): return jsonify({'error':'Unauthorized'}), 403
    d = request.json or {}
    inst_id = d.get('Inst_ID')
    amount = float(d.get('Allocated_Amount') or 0.0)
    year = get_fy_prefix(d.get('year') or '')
    if not inst_id: return jsonify({'error':'Institution ID required'}), 400
    
    try:
        db_query("DELETE FROM Store_Budgets WHERE Year1=%s AND Inst_ID=%s", (year, inst_id), fetch=False)
        db_query("INSERT INTO Store_Budgets (Year1, Inst_ID, Allocated_Amount) VALUES (%s, %s, %s)", (year, inst_id, amount), fetch=False)
        log_audit('Budgets', 'SetStore', inst_id, '', str(amount))
        return jsonify({'success': True})
    except Exception as e: return jsonify({'error': str(e)}), 500

@app.route('/api/budgets/user', methods=['GET'])
def get_user_budgets():
    year = request.args.get('year') or str(datetime.now().year)
    year_prefix = get_fy_prefix(year)
    
    # Store Manager can only view their own budget allocation
    role = session.get('role')
    user_name = session.get('username')
    user_where = ""
    args = [year_prefix]
    if role == 'hostel' and user_name:
        user_where = "WHERE u.username = %s"
        args.append(user_name)

    users = db_query(f"""
        SELECT u.username, u.name, u.role, u.inst_id, COALESCE(ub.Allocated_Amount, 0.0) AS Allocated_Amount
        FROM Users u
        LEFT JOIN User_Budgets ub ON u.username = ub.Username AND ub.Year1 = %s
        {user_where}
        ORDER BY u.username
    """, tuple(args))
    
    # Spends
    grocery_spends = db_query("""
        SELECT Issue_Inst_ID, SUM(Issue_Amount) AS spent
        FROM Stock_Issue
        WHERE Purchased_Donation = 'Issue' AND Year1 = %s AND Issue_Inst_ID IS NOT NULL
        GROUP BY Issue_Inst_ID
    """, (year_prefix,))
    g_spend_map = {row['Issue_Inst_ID']: row['spent'] for row in grocery_spends}
    
    veg_spends = db_query("""
        SELECT Inst_ID, SUM(Quantity * Rate) AS spent
        FROM Vegetable
        WHERE Year1 = %s AND Inst_ID IS NOT NULL
        GROUP BY Inst_ID
    """, (year_prefix,))
    v_spend_map = {row['Inst_ID']: row['spent'] for row in veg_spends}
    
    for u in users:
        iid = u['inst_id']
        if iid:
            spent = g_spend_map.get(iid, 0.0) + v_spend_map.get(iid, 0.0)
        else:
            spent = 0.0
        u['Used_Amount'] = spent
        u['Remaining_Amount'] = u['Allocated_Amount'] - spent
        
    return jsonify(users)

@app.route('/api/budgets/user', methods=['POST'])
def set_user_budget():
    if not chk('head'): return jsonify({'error':'Unauthorized'}), 403
    d = request.json or {}
    username = d.get('Username')
    amount = float(d.get('Allocated_Amount') or 0.0)
    year = get_fy_prefix(d.get('year') or '')
    if not username: return jsonify({'error':'Username required'}), 400
    
    try:
        db_query("DELETE FROM User_Budgets WHERE Year1=%s AND Username=%s", (year, username), fetch=False)
        db_query("INSERT INTO User_Budgets (Year1, Username, Allocated_Amount) VALUES (%s, %s, %s)", (year, username, amount), fetch=False)
        log_audit('Budgets', 'SetUser', username, '', str(amount))
        return jsonify({'success': True})
    except Exception as e: return jsonify({'error': str(e)}), 500


# --- ANALYTICS DASHBOARD ---

@app.route('/api/analytics', methods=['GET'])
def get_analytics():
    year = request.args.get('year') or str(datetime.now().year)
    year_prefix = get_fy_prefix(year)
    role = session.get('role', 'hostel')
    inst_id = session.get('inst_id')
    
    # 1. Central Stock Value
    val_row = db_query("SELECT SUM(Tot_Stock * Std_Rate) AS val FROM Grocery_Items")
    central_stock_val = val_row[0]['val'] or 0.0 if val_row else 0.0
    
    # 2. Store stock value
    store_stock_vals = {}
    for iid, col in INST_ID_TO_COLUMN.items():
        if role == 'hostel' and inst_id != iid:
            continue
        r = db_query(f"SELECT SUM({col} * Std_Rate) AS val FROM Grocery_Items WHERE {col} > 0")
        store_stock_vals[iid] = r[0]['val'] or 0.0 if r else 0.0

    # 3. Budget totals (dynamically calculated from item-wise budget columns * Std_Rate)
    budget_summary = {}
    if role == 'head' or role == 'accounts':
        # Sum of sanctioned budget across all items and all stores
        total_allocated_row = db_query("SELECT SUM(Tot_Quantity * Std_Rate) AS val FROM Grocery_Items")
        total_allocated = total_allocated_row[0]['val'] or 0.0 if total_allocated_row else 0.0
        
        spent_g = db_query("SELECT SUM(Issue_Amount) AS val FROM Stock_Issue WHERE Purchased_Donation='Issue' AND Year1 LIKE %s", (year_prefix + '%',))
        spent_v = db_query("SELECT SUM(Quantity * Rate) AS val FROM Vegetable WHERE Year1 LIKE %s", (year_prefix + '%',))
        total_spent = (spent_g[0]['val'] or 0.0) + (spent_v[0]['val'] or 0.0)
        
        budget_summary = {
            'Total_Budget': total_allocated,
            'Total_Utilized': total_spent,
            'Remaining_Budget': total_allocated - total_spent
        }
    elif inst_id:
        # Sum of sanctioned budget for this store
        budget_col = INST_ID_TO_BUDGET_COLUMN.get(int(inst_id))
        allocated = 0.0
        if budget_col:
            allocated_row = db_query(f"SELECT SUM({budget_col} * Std_Rate) AS val FROM Grocery_Items WHERE {budget_col} > 0")
            allocated = allocated_row[0]['val'] or 0.0 if allocated_row else 0.0
            
        spent_g = db_query("SELECT SUM(Issue_Amount) AS val FROM Stock_Issue WHERE Purchased_Donation='Issue' AND Year1 LIKE %s AND Issue_Inst_ID=%s", (year_prefix + '%', inst_id))
        spent_v = db_query("SELECT SUM(Quantity * Rate) AS val FROM Vegetable WHERE Year1 LIKE %s AND Inst_ID=%s", (year_prefix + '%', inst_id))
        spent = (spent_g[0]['val'] or 0.0) + (spent_v[0]['val'] or 0.0)
        
        budget_summary = {
            'Total_Budget': allocated,
            'Total_Utilized': spent,
            'Remaining_Budget': allocated - spent
        }
        
    # 4. Monthly Spending Graph (calculate in python for dual db compatibility)
    if role == 'head' or role == 'accounts':
        all_issues = db_query("SELECT Date1, Issue_Amount FROM Stock_Issue WHERE Purchased_Donation='Issue' AND Year1=%s", (year_prefix,))
        all_vegs = db_query("SELECT Purchase_On AS Date1, (Quantity * Rate) AS Issue_Amount FROM Vegetable WHERE Year1=%s", (year_prefix,))
    else:
        all_issues = db_query("SELECT Date1, Issue_Amount FROM Stock_Issue WHERE Purchased_Donation='Issue' AND Year1=%s AND Issue_Inst_ID=%s", (year_prefix, inst_id))
        all_vegs = db_query("SELECT Purchase_On AS Date1, (Quantity * Rate) AS Issue_Amount FROM Vegetable WHERE Year1=%s AND Inst_ID=%s", (year_prefix, inst_id))
        
    month_totals = {f"{m:02d}": 0.0 for m in range(1, 13)}
    for row in all_issues + all_vegs:
        dt = row['Date1']
        if dt and len(dt) >= 7:
            m = dt[5:7]
            if m in month_totals:
                month_totals[m] += row['Issue_Amount'] or 0.0
                
    monthly_spend = [{'Month': m, 'Amount': amt} for m, amt in sorted(month_totals.items())]

    # 5. Consumption Analytics (Item wise)
    if role == 'head' or role == 'accounts':
        top_consumed = db_query("""
            SELECT gi.Grocery_Items_Kan, SUM(si.Issue) AS TotalQty
            FROM Stock_Issue si JOIN Grocery_Items gi ON si.Grocery_Code = gi.Grocery_Code
            WHERE si.Purchased_Donation = 'Consumption' AND si.Year1 = %s
            GROUP BY gi.Grocery_Code, gi.Grocery_Items_Kan
            ORDER BY TotalQty DESC LIMIT 10
        """, (year_prefix,))
    else:
        top_consumed = db_query("""
            SELECT gi.Grocery_Items_Kan, SUM(si.Issue) AS TotalQty
            FROM Stock_Issue si JOIN Grocery_Items gi ON si.Grocery_Code = gi.Grocery_Code
            WHERE si.Purchased_Donation = 'Consumption' AND si.Year1 = %s AND si.Issue_Inst_ID = %s
            GROUP BY gi.Grocery_Code, gi.Grocery_Items_Kan
            ORDER BY TotalQty DESC LIMIT 10
        """, (year_prefix, inst_id))

    # 6. Fast/Slow Central Items (Admin Only)
    fast_moving = []
    slow_moving = []
    if role == 'head':
        fast_moving = db_query("""
            SELECT gi.Grocery_Items_Kan, SUM(si.Issue) AS TotalQty
            FROM Stock_Issue si JOIN Grocery_Items gi ON si.Grocery_Code = gi.Grocery_Code
            WHERE si.Purchased_Donation = 'Issue' AND si.Year1 = %s
            GROUP BY gi.Grocery_Code, gi.Grocery_Items_Kan
            ORDER BY TotalQty DESC LIMIT 5
        """, (year_prefix,))
        
        slow_moving = db_query("""
            SELECT gi.Grocery_Items_Kan, gi.Tot_Stock
            FROM Grocery_Items gi
            WHERE gi.Tot_Stock > 0 AND gi.Grocery_Code NOT IN (
                SELECT DISTINCT Grocery_Code FROM Stock_Issue WHERE Purchased_Donation='Issue' AND Year1=%s
            )
            ORDER BY gi.Tot_Stock DESC LIMIT 5
        """, (year_prefix,))

    # 7. User Action stats (Admin Only)
    user_stats = []
    if role == 'head':
        user_stats = db_query("""
            SELECT Username, Action, COUNT(*) AS ActionCount
            FROM Audit_Logs
            GROUP BY Username, Action
            ORDER BY ActionCount DESC LIMIT 30
        """)

    return jsonify({
        'Central_Stock_Value': central_stock_val,
        'Store_Stock_Values': store_stock_vals,
        'Budget_Summary': budget_summary,
        'Monthly_Spend': monthly_spend,
        'Top_Consumed': top_consumed,
        'Fast_Moving': fast_moving,
        'Slow_Moving': slow_moving,
        'User_Stats': user_stats
    })


# --- REPORTS ---

@app.route('/api/reports/stock', methods=['GET'])
def report_stock():
    cat = request.args.get('category','')
    year = request.args.get('year') or str(datetime.now().year)
    year_prefix = get_fy_prefix(year)
    
    wheres = ["gi.Grocery_Items_Kan != 'x'", "gi.Grocery_Items_Kan != ''", "gi.Grocery_Items_Kan IS NOT NULL"]
    args = [year_prefix + '%', year_prefix + '%']
    if cat:
        wheres.append("gi.Grocery_Category=%s")
        args.append(cat)
        
    where_clause = "WHERE " + " AND ".join(wheres)
    
    query = f"""
        SELECT 
            gi.Grocery_Code,
            gi.Grocery_Items_Kan,
            gi.Grocery_Items_Eng,
            gi.Grocery_Category,
            gi.Qtl_Kg_Ltr,
            gi.Std_Rate,
            COALESCE(inward.tot, 0.0) AS Tot_Stock,
            COALESCE(outward.tot, 0.0) AS Tot_Issue,
            (COALESCE(inward.tot, 0.0) * gi.Std_Rate) AS StockValue
        FROM Grocery_Items gi
        LEFT JOIN (
            SELECT Grocery_Code, SUM(Stock) AS tot 
            FROM Stock_Issue 
            WHERE Year1 LIKE %s AND Stock > 0 
            GROUP BY Grocery_Code
        ) inward ON gi.Grocery_Code = inward.Grocery_Code
        LEFT JOIN (
            SELECT Grocery_Code, SUM(Issue) AS tot 
            FROM Stock_Issue 
            WHERE Year1 LIKE %s AND Issue > 0 
            GROUP BY Grocery_Code
        ) outward ON gi.Grocery_Code = outward.Grocery_Code
        {where_clause}
        ORDER BY gi.Grocery_Category, gi.Grocery_Code
    """
    return jsonify(db_query(query, tuple(args)))

@app.route('/api/reports/purchases', methods=['GET'])
def report_purchases():
    df = request.args.get('from',''); dt = request.args.get('to',''); dn = request.args.get('donor',''); tp = request.args.get('type','')
    fy = request.args.get('year', '')
    wheres = ["si.Stock>0"]; args = []
    if df: wheres.append("si.Date1>=%s"); args.append(df)
    if dt: wheres.append("si.Date1<=%s"); args.append(dt)
    if dn: wheres.append("si.Shop_Donor_ID=%s"); args.append(dn)
    if tp: wheres.append("si.Purchased_Donation=%s"); args.append(tp)
    if fy: wheres.append("si.Year1 LIKE %s"); args.append(get_fy_prefix(fy) + "%")
    
    return jsonify(db_query(f"""SELECT si.Date1,si.Grocery_Code,gi.Grocery_Items_Kan,gi.Grocery_Items_Eng,
        gi.Qtl_Kg_Ltr,si.Stock AS Quantity,si.Purchase_Rate,(si.Stock*si.Purchase_Rate) AS Amount,
        sd.Shop_Donor_Name,si.Bill_No,si.Purchased_Donation,si.Remarks
        FROM Stock_Issue si JOIN Grocery_Items gi ON si.Grocery_Code=gi.Grocery_Code
        LEFT JOIN Shops_Donors sd ON si.Shop_Donor_ID=sd.Shop_Donor_ID
        WHERE {' AND '.join(wheres)} ORDER BY si.Date1 DESC,si.Rec DESC""", tuple(args)))

@app.route('/api/reports/indents', methods=['GET'])
def report_indents():
    df = request.args.get('from',''); dt = request.args.get('to',''); sf = request.args.get('status',''); iid = request.args.get('inst_id','')
    fy = request.args.get('year', '')
    wheres = []; args = []
    if df: wheres.append("ind.Indent_Date>=%s"); args.append(df)
    if dt: wheres.append("ind.Indent_Date<=%s"); args.append(dt)
    if sf: wheres.append("ind.Sanctioned=%s"); args.append(sf)
    if iid: wheres.append("ind.Inst_ID=%s"); args.append(iid)
    if fy: wheres.append("ind.Year1 LIKE %s"); args.append(get_fy_prefix(fy) + "%")
    
    where_sql = ("WHERE " + " AND ".join(wheres)) if wheres else ""
    return jsonify(db_query(f"""SELECT ind.Indent_Date,ind.Indent_no,inst.Institution,
        gi.Grocery_Items_Kan,gi.Grocery_Items_Eng,gi.Qtl_Kg_Ltr,
        ind.Quantity,ind.Sanctioned_Quantity,ind.Sanctioned,ind.Sanctioned_on,ind.Remarks
        FROM Indents ind JOIN Grocery_Items gi ON ind.Grocery_Code=gi.Grocery_Code
        JOIN Institutions inst ON ind.Inst_ID=inst.Inst_ID {where_sql}
        ORDER BY ind.Indent_Date DESC,ind.Rec DESC""", tuple(args)))

@app.route('/api/reports/consumption', methods=['GET'])
def report_consumption():
    df = request.args.get('from',''); dt = request.args.get('to',''); iid = request.args.get('inst_id','')
    item_code = request.args.get('item_code','')
    fy = request.args.get('year', '')
    wheres = ["si.Purchased_Donation='Consumption'"]; args = []
    if df: wheres.append("si.Date1>=%s"); args.append(df)
    if dt: wheres.append("si.Date1<=%s"); args.append(dt)
    if iid: wheres.append("si.Issue_Inst_ID=%s"); args.append(iid)
    if item_code: wheres.append("si.Grocery_Code=%s"); args.append(item_code)
    if fy: wheres.append("si.Year1 LIKE %s"); args.append(get_fy_prefix(fy) + "%")
    
    logs = db_query(f"""SELECT si.Date1,inst.Institution,gi.Grocery_Code,gi.Grocery_Items_Kan,gi.Grocery_Items_Eng,
        gi.Qtl_Kg_Ltr,gi.Grocery_Category,si.Issue AS QuantityUsed,si.Issue_Amount,si.Remarks,
        gi.Boys_Hostel_Qty, gi.Girls_Hostel_Qty, gi.Math_Qty, gi.Shantivan_Qty_a, gi.Hunnime_Qty, gi.AO_Office_Qty, gi.Shraddanjali_Qty,
        si.Issue_Inst_ID
        FROM Stock_Issue si JOIN Grocery_Items gi ON si.Grocery_Code=gi.Grocery_Code
        JOIN Institutions inst ON si.Issue_Inst_ID=inst.Inst_ID
        WHERE {' AND '.join(wheres)} ORDER BY si.Date1 DESC,si.Rec DESC""", tuple(args))
        
    for log in logs:
        inst_id = int(log['Issue_Inst_ID'])
        col = INST_ID_TO_COLUMN.get(inst_id)
        if col and col in log:
            log['RemainingStock'] = log[col]
        else:
            log['RemainingStock'] = 0.0
            
    return jsonify(logs)

@app.route('/api/reports/donations', methods=['GET'])
def report_donations():
    df = request.args.get('from',''); dt = request.args.get('to',''); dn = request.args.get('donor','')
    fy = request.args.get('year', '')
    wheres = ["si.Purchased_Donation='Donation'"]; args = []
    if df: wheres.append("si.Date1>=%s"); args.append(df)
    if dt: wheres.append("si.Date1<=%s"); args.append(dt)
    if dn: wheres.append("si.Shop_Donor_ID=%s"); args.append(dn)
    if fy: wheres.append("si.Year1 LIKE %s"); args.append(get_fy_prefix(fy) + "%")
    
    return jsonify(db_query(f"""SELECT si.Date1,sd.Shop_Donor_Name,sd.Place,
        gi.Grocery_Items_Kan,gi.Grocery_Items_Eng,gi.Qtl_Kg_Ltr,si.Stock AS Quantity,si.Remarks
        FROM Stock_Issue si JOIN Grocery_Items gi ON si.Grocery_Code=gi.Grocery_Code
        LEFT JOIN Shops_Donors sd ON si.Shop_Donor_ID=sd.Shop_Donor_ID
        WHERE {' AND '.join(wheres)} ORDER BY si.Date1 DESC""", tuple(args)))


# --- AUDIT LOGS ---

@app.route('/api/audit-logs', methods=['GET'])
def get_audit_logs():
    if not chk('head'): return jsonify({'error':'Unauthorized'}), 403
    return jsonify(db_query("SELECT * FROM Audit_Logs ORDER BY Rec DESC LIMIT 500"))


# --- USERS ---

@app.route('/api/users', methods=['GET'])
def get_users():
    if not chk('head'): return jsonify({'error':'Unauthorized'}), 403
    return jsonify(db_query("SELECT username,role,inst_id,name FROM Users ORDER BY role,username"))

@app.route('/api/users', methods=['POST'])
def create_user():
    if not chk('head'): return jsonify({'error':'Unauthorized'}), 403
    d = request.json or {}
    un = d.get('username','').strip().lower(); pw = d.get('password','').strip()
    role = d.get('role','').strip(); name = d.get('name','').strip()
    if not un or not pw or not role or not name: return jsonify({'error':'All fields required'}), 400
    try:
        if db_query("SELECT username FROM Users WHERE username=%s", (un,)): return jsonify({'error':'Username exists'}), 400
        db_query("INSERT INTO Users (username,password,role,inst_id,name) VALUES (%s,%s,%s,%s,%s)",
            (un, pw, role, d.get('inst_id'), name), fetch=False)
        log_audit('Users','Create',un,'',role)
        return jsonify({'success': True})
    except Exception as e: return jsonify({'error': str(e)}), 500

@app.route('/api/users/<username>', methods=['PUT'])
def edit_user(username):
    if not chk('head'): return jsonify({'error':'Unauthorized'}), 403
    d = request.json or {}
    name = d.get('name','').strip(); role = d.get('role','').strip()
    pw = d.get('password','').strip(); inst_id = d.get('inst_id')
    try:
        if pw:
            db_query("UPDATE Users SET name=%s,role=%s,inst_id=%s,password=%s WHERE username=%s", (name,role,inst_id,pw,username), fetch=False)
        else:
            db_query("UPDATE Users SET name=%s,role=%s,inst_id=%s WHERE username=%s", (name,role,inst_id,username), fetch=False)
        log_audit('Users','Edit',username,'',role)
        return jsonify({'success': True})
    except Exception as e: return jsonify({'error': str(e)}), 500

@app.route('/api/users/<username>', methods=['DELETE'])
def delete_user(username):
    if not chk('head'): return jsonify({'error':'Unauthorized'}), 403
    if username == session.get('username'): return jsonify({'error':'Cannot delete own account'}), 400
    try:
        db_query("DELETE FROM Users WHERE username=%s", (username,), fetch=False)
        log_audit('Users','Delete',username)
        return jsonify({'success': True})
    except Exception as e: return jsonify({'error': str(e)}), 500


# --- DB INIT ---

def migrate_and_load_budgets():
    try:
        conn, db_type = get_db_connection()
        cursor = conn.cursor()
        
        # 1. Clean up dummy 'x' items
        print("Cleaning up dummy 'x' items from Grocery_Items...")
        if db_type == "postgresql":
            cursor.execute("DELETE FROM Grocery_Items WHERE Grocery_Items_Kan = 'x' OR Grocery_Items_Kan = '' OR Grocery_Items_Kan IS NULL")
        else:
            cursor.execute("DELETE FROM Grocery_Items WHERE Grocery_Items_Kan = 'x' OR Grocery_Items_Kan = '' OR Grocery_Items_Kan IS NULL")
        
        # 2. Alter table to add budget columns if they don't exist
        budget_cols = [
            "Shraddanjali_Budget", "Hunnime_Budget", "Boys_Hostel_Budget",
            "Girls_Hostel_Budget", "Math_Budget", "Shantivan_Budget_a", "AO_Office_Budget"
        ]
        
        if db_type == "postgresql":
            cursor.execute("SELECT column_name FROM information_schema.columns WHERE table_name = 'grocery_items'")
            existing_cols = {row[0].lower() for row in cursor.fetchall()}
        else:
            cursor.execute("PRAGMA table_info(Grocery_Items)")
            existing_cols = {row[1].lower() for row in cursor.fetchall()}
            
        for col in budget_cols:
            if col.lower() not in existing_cols:
                print(f"Altering table: Adding column {col} to Grocery_Items...")
                cursor.execute(f"ALTER TABLE Grocery_Items ADD COLUMN {col} REAL DEFAULT 0.0")
        
        conn.commit()

        # 3. Update budget quantities from Excel sheet if available
        base_dir = os.path.dirname(os.path.abspath(__file__))
        xlsx_path = os.path.join(base_dir, 'Godown & Store.xlsx')
        if os.path.exists(xlsx_path):
            import openpyxl
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
            sheet = wb['Grocery_Items'] if 'Grocery_Items' in wb.sheetnames else wb.sheetnames[0]
            rows = list(sheet.iter_rows(values_only=True))
            if rows:
                headers = [str(cell).strip() for cell in rows[0] if cell is not None]
                h_map = {h.lower().replace(" ", "_"): idx for idx, h in enumerate(headers)}
                
                code_idx = h_map.get('grocery_code') or h_map.get('code')
                if code_idx is not None:
                    def get_val(row_data, keys, default=0.0):
                        for k in keys:
                            k_mod = k.lower().replace(" ", "_")
                            if k_mod in h_map:
                                val = row_data[h_map[k_mod]]
                                try: return float(val or default)
                                except: pass
                        return default
                    
                    updates = 0
                    for row in rows[1:]:
                        if not row or all(c is None for c in row):
                            continue
                        try:
                            code = int(float(row[code_idx]))
                        except:
                            continue
                        
                        name_kan = str(row[h_map.get('grocery_items') or h_map.get('grocery_items_kan') or 2] or '')
                        if name_kan.lower().strip() == 'x' or not name_kan.strip():
                            continue
                            
                        shraddanjali = get_val(row, ['shraddanjali', 'store_-_shraddhajali'])
                        hunnime = get_val(row, ['hunnime', 'stores_-_a'])
                        boys = get_val(row, ['boys_hostel', 'stores_-_boys_hostel', 'stores_-_sirigere_bhs'])
                        girls = get_val(row, ['girls_hostel', 'stores_-_girls_hostel', 'stores_-_sirigere_ghs'])
                        math_val = get_val(row, ['math', 'stores_-_math'])
                        shantivan = get_val(row, ['shantivana', 'stores_-_shantivana_bidara', 'stores_-_shantivana_gurukula'])
                        ao_office = get_val(row, ['ao_office', 'stores_-_ao_office'])
                        
                        update_query = """
                            UPDATE Grocery_Items
                            SET Shraddanjali_Budget = %s, Hunnime_Budget = %s, Boys_Hostel_Budget = %s,
                                Girls_Hostel_Budget = %s, Math_Budget = %s, Shantivan_Budget_a = %s, AO_Office_Budget = %s
                            WHERE Grocery_Code = %s
                        """ if db_type == "postgresql" else """
                            UPDATE Grocery_Items
                            SET Shraddanjali_Budget = ?, Hunnime_Budget = ?, Boys_Hostel_Budget = ?,
                                Girls_Hostel_Budget = ?, Math_Budget = ?, Shantivan_Budget_a = ?, AO_Office_Budget = ?
                            WHERE Grocery_Code = ?
                        """
                        cursor.execute(update_query, (shraddanjali, hunnime, boys, girls, math_val, shantivan, ao_office, code))
                        updates += 1
                    conn.commit()
                    print(f"Successfully loaded and updated budget details for {updates} items from Excel.")
        
        conn.close()
    except Exception as ex:
        print(f"Error in migrate_and_load_budgets: {ex}")

def init_db():
    try:
        db_query("CREATE TABLE IF NOT EXISTS Users (username VARCHAR(50) PRIMARY KEY, password VARCHAR(100) NOT NULL, role VARCHAR(20) NOT NULL, inst_id INT, name VARCHAR(100) NOT NULL)", fetch=False)
        
        # Determine database connection type to run appropriate autoincrement commands
        conn, db_type = get_db_connection()
        conn.close()
        
        if db_type == "sqlite":
            db_query("CREATE TABLE IF NOT EXISTS Audit_Logs (Rec INTEGER PRIMARY KEY AUTOINCREMENT, Timestamp TEXT NOT NULL, Username TEXT NOT NULL, Module TEXT NOT NULL, Action TEXT NOT NULL, Target_ID TEXT, Old_Value TEXT, New_Value TEXT)", fetch=False)
            db_query("CREATE TABLE IF NOT EXISTS Store_Budgets (Rec INTEGER PRIMARY KEY AUTOINCREMENT, Year1 TEXT NOT NULL, Inst_ID INTEGER NOT NULL, Allocated_Amount REAL DEFAULT 0.0, FOREIGN KEY(Inst_ID) REFERENCES Institutions(Inst_ID))", fetch=False)
            db_query("CREATE TABLE IF NOT EXISTS User_Budgets (Rec INTEGER PRIMARY KEY AUTOINCREMENT, Year1 TEXT NOT NULL, Username TEXT NOT NULL, Allocated_Amount REAL DEFAULT 0.0, FOREIGN KEY(Username) REFERENCES Users(username))", fetch=False)
        else:
            db_query("CREATE TABLE IF NOT EXISTS Audit_Logs (Rec SERIAL PRIMARY KEY, Timestamp VARCHAR(50) NOT NULL, Username VARCHAR(100) NOT NULL, Module VARCHAR(100) NOT NULL, Action VARCHAR(100) NOT NULL, Target_ID VARCHAR(100), Old_Value TEXT, New_Value TEXT)", fetch=False)
            db_query("CREATE TABLE IF NOT EXISTS Store_Budgets (Rec SERIAL PRIMARY KEY, Year1 VARCHAR(10) NOT NULL, Inst_ID INTEGER NOT NULL, Allocated_Amount REAL DEFAULT 0.0, FOREIGN KEY(Inst_ID) REFERENCES Institutions(Inst_ID))", fetch=False)
            db_query("CREATE TABLE IF NOT EXISTS User_Budgets (Rec SERIAL PRIMARY KEY, Year1 VARCHAR(10) NOT NULL, Username VARCHAR(50) NOT NULL, Allocated_Amount REAL DEFAULT 0.0, FOREIGN KEY(Username) REFERENCES Users(username))", fetch=False)
            
        existing = db_query("SELECT COUNT(*) as cnt FROM Users")
        count = existing[0].get('cnt', 0) if existing else 0
        if count == 0:
            for un, info in USERS.items():
                db_query("INSERT INTO Users (username,password,role,inst_id,name) VALUES (%s,%s,%s,%s,%s)",
                    (un, info['password'], info['role'], info['inst_id'], info['name']), fetch=False)
            print("Default users seeded.")
        
        # Run migrations and seed budgets dynamically
        migrate_and_load_budgets()
        
        # Check and load all register records if empty
        check_and_load_registers()
        
    except Exception as e:
        print(f"DB init error: {e}")

def check_and_load_registers():
    try:
        # Check if already loaded
        existing = db_query("SELECT COUNT(*) as cnt FROM Stock_Issue")
        count = existing[0]['cnt'] if existing else 0
        if count >= 100:
            print(f"Historical registers already loaded ({count} issues found). Skipping initial import.")
            return
            
        print("Stock_Issue table is empty. Loading all historical Excel registers into database...")
        conn, db_type = get_db_connection()
        cursor = conn.cursor()
        
        # Clear any partial loads to ensure clean seeding
        cursor.execute("DELETE FROM Shops_Donors")
        cursor.execute("DELETE FROM Bills")
        cursor.execute("DELETE FROM Indents")
        cursor.execute("DELETE FROM Stock_Issue")
        cursor.execute("DELETE FROM Vegetable")
        conn.commit()

        base_dir = os.path.dirname(os.path.abspath(__file__))
        files_map = {
            '2025': 'Godown Register 2025-26.xlsx',
            '2026': 'Godown Register 2026-27.xlsx'
        }

        placeholder_donor_id = 999999
        q_donor_init = "INSERT INTO Shops_Donors (Year1, Shop_Donor_ID, Shop_Donor_Name, DateStamp) VALUES (%s,%s,%s,%s) ON CONFLICT DO NOTHING" if db_type == "postgresql" else "INSERT OR IGNORE INTO Shops_Donors (Year1, Shop_Donor_ID, Shop_Donor_Name, DateStamp) VALUES (?,?,?,?)"
        cursor.execute(q_donor_init, ('2025', placeholder_donor_id, 'Fallback General Donor', '2026-07-06'))
        conn.commit()

        # Define inline helper functions
        def format_date(val):
            if val is None: return ""
            import datetime
            if isinstance(val, (datetime.datetime, datetime.date)): return val.strftime('%Y-%m-%d')
            s = str(val).strip()
            if ' ' in s: s = s.split(' ')[0]
            return s

        def clean_float(val, default=0.0):
            if val is None: return default
            try: return float(val)
            except: return default

        def clean_int(val, default=0):
            if val is None: return default
            try: return int(float(val))
            except: return default

        import openpyxl
        import datetime

        for year_prefix, filename in files_map.items():
            filepath = os.path.join(base_dir, filename)
            if not os.path.exists(filepath):
                print(f"Excel Register file {filename} not found, skipping...")
                continue
                
            print(f"Importing {filename} for year {year_prefix}...")
            wb = openpyxl.load_workbook(filepath, data_only=True)
            
            # 1. Shops_Donors
            if 'Shops_Donors' in wb.sheetnames:
                ws = wb['Shops_Donors']
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) > 1:
                    headers = [str(h).strip().lower() for h in rows[0]]
                    h_map = {h: i for i, h in enumerate(headers)}
                    donors = []
                    for r in rows[1:]:
                        if not r or all(c is None for c in r): continue
                        did = clean_int(r[h_map.get('shop_donor_id', 0)])
                        name = str(r[h_map.get('name', 1)] or '').strip()
                        if not did or not name: continue
                        pid = clean_int(r[h_map.get('place_id')]) if 'place_id' in h_map else 0
                        place = str(r[h_map.get('place')] or '') if 'place' in h_map else ''
                        taluk = str(r[h_map.get('taluk')] or '') if 'taluk' in h_map else ''
                        dist = str(r[h_map.get('dist')] or '') if 'dist' in h_map else ''
                        pin = str(r[h_map.get('pin')] or '') if 'pin' in h_map else ''
                        mob = str(r[h_map.get('mobile')] or '') if 'mobile' in h_map else ''
                        rem = str(r[h_map.get('remarks')] or '') if 'remarks' in h_map else ''
                        donors.append((
                            year_prefix, did, name, pid, place, place, taluk, taluk, dist, dist, pin, 'Karnataka', 'India', mob, rem, datetime.date.today().strftime('%Y-%m-%d')
                        ))
                    if donors:
                        q = '''
                            INSERT INTO Shops_Donors 
                            (Year1, Shop_Donor_ID, Shop_Donor_Name, Place_ID, Place, Place_Kan, Taluk, Taluk_Kan, Dist, Dist_Kan, Pin, State, Country, Mobile, Remarks, DateStamp)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) ON CONFLICT (Shop_Donor_ID) DO NOTHING
                        ''' if db_type == "postgresql" else '''
                            INSERT OR IGNORE INTO Shops_Donors 
                            (Year1, Shop_Donor_ID, Shop_Donor_Name, Place_ID, Place, Place_Kan, Taluk, Taluk_Kan, Dist, Dist_Kan, Pin, State, Country, Mobile, Remarks, DateStamp)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                        '''
                        cursor.executemany(q, donors)
                        conn.commit()

            # 2. Bills
            if 'Bills' in wb.sheetnames:
                ws = wb['Bills']
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) > 1:
                    headers = [str(h).strip().lower() for h in rows[0]]
                    h_map = {h: i for i, h in enumerate(headers)}
                    bills = []
                    for r in rows[1:]:
                        if not r or all(c is None for c in r): continue
                        shop_id = clean_int(r[h_map.get('shop_id', 0)])
                        bill_date = format_date(r[h_map.get('bill_date', 2)])
                        bill_no = str(r[h_map.get('bill_no', 3)] or '').strip()
                        bill_amt = clean_float(r[h_map.get('bill_amount', 4)])
                        if not shop_id or not bill_date or not bill_no: continue
                        paid_by = str(r[h_map.get('paid_by')] or '') if 'paid_by' in h_map else ''
                        ch_date = format_date(r[h_map.get('ch_date')]) if 'ch_date' in h_map else ''
                        ch_no = str(r[h_map.get('ch_no') or '']) if 'ch_no' in h_map else ''
                        ch_amt = clean_float(r[h_map.get('ch_amount')]) if 'ch_amount' in h_map else 0.0
                        rem = str(r[h_map.get('remarks')] or '') if 'remarks' in h_map else ''
                        bills.append((
                            year_prefix, shop_id, bill_date, bill_no, bill_amt, paid_by, ch_date, ch_no, ch_amt, rem, datetime.date.today().strftime('%Y-%m-%d')
                        ))
                    if bills:
                        q = '''
                            INSERT INTO Bills (Year1, Shop_Donor_ID, Bill_Date, Bill_No, Bill_Amount, Paid_By, Ch_Date, Ch_No, Ch_Amount, Remarks, DateAdded)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        ''' if db_type == "postgresql" else '''
                            INSERT INTO Bills (Year1, Shop_Donor_ID, Bill_Date, Bill_No, Bill_Amount, Paid_By, Ch_Date, Ch_No, Ch_Amount, Remarks, DateAdded)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?)
                        '''
                        cursor.executemany(q, bills)
                        conn.commit()

            # 3. Indents
            if 'Indents' in wb.sheetnames:
                ws = wb['Indents']
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) > 1:
                    headers = [str(h).strip().lower().replace(" ", "_") for h in rows[0]]
                    h_map = {h: i for i, h in enumerate(headers)}
                    indents = []
                    for r in rows[1:]:
                        if not r or all(c is None for c in r): continue
                        code = clean_int(r[h_map.get('code', 3)])
                        inst_id = clean_int(r[h_map.get('inst_id', 1)])
                        qty = clean_float(r[h_map.get('quantity', 5)])
                        indent_no = str(r[h_map.get('indent_no', 9)] or '').strip()
                        if not code or not inst_id or not qty or not indent_no: continue
                        indent_date = format_date(r[h_map.get('indent_date', 0)])
                        sanc_qty = clean_float(r[h_map.get('sanctioned_quantity', 10)])
                        status = str(r[h_map.get('sanctioned', 11)] or 'Pending').strip()
                        sanc_on = format_date(r[h_map.get('sanctioned_on')]) if 'sanctioned_on' in h_map else ''
                        indents.append((
                            year_prefix, indent_date, code, inst_id, qty, indent_no, sanc_qty, status, sanc_on, datetime.date.today().strftime('%Y-%m-%d'), 'Imported'
                        ))
                    if indents:
                        q = '''
                            INSERT INTO Indents (Year1, Indent_Date, Grocery_Code, Inst_ID, Quantity, Indent_no, Sanctioned_Quantity, Sanctioned, Sanctioned_on, DateStamp, Remarks)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        ''' if db_type == "postgresql" else '''
                            INSERT INTO Indents (Year1, Indent_Date, Grocery_Code, Inst_ID, Quantity, Indent_no, Sanctioned_Quantity, Sanctioned, Sanctioned_on, DateStamp, Remarks)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?)
                        '''
                        cursor.executemany(q, indents)
                        conn.commit()
                        
            # 3.5 Godown (Opening Stock)
            if 'Godown' in wb.sheetnames:
                ws = wb['Godown']
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) > 1:
                    headers = [str(h).strip().lower().replace(" ", "_") for h in rows[0]]
                    h_map = {h: i for i, h in enumerate(headers)}
                    openings = []
                    opening_date = f"{year_prefix}-04-01"
                    for r in rows[1:]:
                        if not r or all(c is None for c in r): continue
                        code = clean_int(r[h_map.get('grocery_code', 1)])
                        open_qty = clean_float(r[h_map.get('opening_stock', 4)])
                        if not code or open_qty <= 0: continue
                        rate = clean_float(r[h_map.get('std_rate', 9)])
                        amt = open_qty * rate
                        openings.append((
                            year_prefix, opening_date, code, placeholder_donor_id, '', '', '', '', '', rate, amt, 'Yes', open_qty, None, 0.0, 0.0, 'Central Godown', datetime.date.today().strftime('%Y-%m-%d'), 'Imported Opening Stock', 'Opening Stock'
                        ))
                    if openings:
                        q = '''
                            INSERT INTO Stock_Issue (Year1, Date1, Grocery_Code, Shop_Donor_ID, Book_No, Receipt_No, Receipt_Date, Bill_Date, Bill_No, Purchase_Rate, Purchase_Amount, Paid, Stock, Issue_Inst_ID, Issue, Issue_Amount, Received_By, DateStamp, Remarks, Purchased_Donation)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        ''' if db_type == "postgresql" else '''
                            INSERT INTO Stock_Issue (Year1, Date1, Grocery_Code, Shop_Donor_ID, Book_No, Receipt_No, Receipt_Date, Bill_Date, Bill_No, Purchase_Rate, Purchase_Amount, Paid, Stock, Issue_Inst_ID, Issue, Issue_Amount, Received_By, DateStamp, Remarks, Purchased_Donation)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                        '''
                        chunk_size = 1000
                        for idx in range(0, len(openings), chunk_size):
                            cursor.executemany(q, openings[idx : idx + chunk_size])
                            conn.commit()

            # 4. Stock (Inwards)
            if 'Stock' in wb.sheetnames:
                ws = wb['Stock']
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) > 1:
                    headers = [str(h).strip().lower().replace(" ", "_") for h in rows[0]]
                    h_map = {h: i for i, h in enumerate(headers)}
                    inwards = []
                    for r in rows[1:]:
                        if not r or all(c is None for c in r): continue
                        code = clean_int(r[h_map.get('grocery_code', 3)])
                        stock_qty = clean_float(r[h_map.get('stock', 5)])
                        if not code or not stock_qty: continue
                        date1 = format_date(r[h_map.get('date', 1)])
                        purch_don = str(r[h_map.get('purchased_donation', 2)] or 'Purchase').strip()
                        rate = clean_float(r[h_map.get('purchase_rate', 7)])
                        amt = clean_float(r[h_map.get('purchase_amount', 9)])
                        book_no = str(r[h_map.get('book_no')] or '') if 'book_no' in h_map else ''
                        receipt_no = str(r[h_map.get('receipt_no')] or '') if 'receipt_no' in h_map else ''
                        receipt_date = format_date(r[h_map.get('receipt_date')]) if 'receipt_date' in h_map else ''
                        
                        donor_id = clean_int(r[h_map.get('shop_donor_id')]) if 'shop_donor_id' in h_map else placeholder_donor_id
                        if not donor_id: donor_id = placeholder_donor_id
                        bill_date = format_date(r[h_map.get('bill_date')]) if 'bill_date' in h_map else ''
                        bill_no = str(r[h_map.get('bill_no')] or '') if 'bill_no' in h_map else ''
                        paid = str(r[h_map.get('paid')] or 'No') if 'paid' in h_map else 'No'
                        received_by = str(r[h_map.get('received_by')] or 'Central Godown') if 'received_by' in h_map else 'Central Godown'
                        
                        inwards.append((
                            year_prefix, date1, code, donor_id, book_no, receipt_no, receipt_date, bill_date, bill_no, rate, amt, paid, stock_qty, None, 0.0, 0.0, received_by, datetime.date.today().strftime('%Y-%m-%d'), 'Imported', purch_don
                        ))
                    if inwards:
                        q = '''
                            INSERT INTO Stock_Issue (Year1, Date1, Grocery_Code, Shop_Donor_ID, Book_No, Receipt_No, Receipt_Date, Bill_Date, Bill_No, Purchase_Rate, Purchase_Amount, Paid, Stock, Issue_Inst_ID, Issue, Issue_Amount, Received_By, DateStamp, Remarks, Purchased_Donation)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        ''' if db_type == "postgresql" else '''
                            INSERT INTO Stock_Issue (Year1, Date1, Grocery_Code, Shop_Donor_ID, Book_No, Receipt_No, Receipt_Date, Bill_Date, Bill_No, Purchase_Rate, Purchase_Amount, Paid, Stock, Issue_Inst_ID, Issue, Issue_Amount, Received_By, DateStamp, Remarks, Purchased_Donation)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                        '''
                        chunk_size = 1000
                        for idx in range(0, len(inwards), chunk_size):
                            cursor.executemany(q, inwards[idx : idx + chunk_size])
                            conn.commit()

            # 5. Issue (Outwards)
            if 'Issue' in wb.sheetnames:
                ws = wb['Issue']
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) > 1:
                    headers = [str(h).strip().lower().replace(" ", "_") for h in rows[0]]
                    h_map = {h: i for i, h in enumerate(headers)}
                    outwards = []
                    for r in rows[1:]:
                        if not r or all(c is None for c in r): continue
                        code = clean_int(r[h_map.get('code', 4)])
                        issue_qty = clean_float(r[h_map.get('issue', 6)])
                        inst_id = clean_int(r[h_map.get('issue_inst_id', 1)])
                        if not code or not issue_qty or not inst_id: continue
                        date1 = format_date(r[h_map.get('date', 3)])
                        rate = clean_float(r[h_map.get('std_rate', 8)])
                        amt = clean_float(r[h_map.get('issue_amt', 9)])
                        outwards.append((
                            year_prefix, date1, code, None, '', '', '', '', '', rate, 0.0, 'No', 0.0, inst_id, issue_qty, amt, 'Receiver', datetime.date.today().strftime('%Y-%m-%d'), 'Imported Outward', 'Issue'
                        ))
                    if outwards:
                        q = '''
                            INSERT INTO Stock_Issue (Year1, Date1, Grocery_Code, Shop_Donor_ID, Book_No, Receipt_No, Receipt_Date, Bill_Date, Bill_No, Purchase_Rate, Purchase_Amount, Paid, Stock, Issue_Inst_ID, Issue, Issue_Amount, Received_By, DateStamp, Remarks, Purchased_Donation)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        ''' if db_type == "postgresql" else '''
                            INSERT INTO Stock_Issue (Year1, Date1, Grocery_Code, Shop_Donor_ID, Book_No, Receipt_No, Receipt_Date, Bill_Date, Bill_No, Purchase_Rate, Purchase_Amount, Paid, Stock, Issue_Inst_ID, Issue, Issue_Amount, Received_By, DateStamp, Remarks, Purchased_Donation)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                        '''
                        chunk_size = 1000
                        for idx in range(0, len(outwards), chunk_size):
                            cursor.executemany(q, outwards[idx : idx + chunk_size])
                            conn.commit()

            # 6. Vegetable
            if 'Vegetable' in wb.sheetnames:
                ws = wb['Vegetable']
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) > 1:
                    headers = [str(h).strip().lower().replace(" ", "_") for h in rows[0]]
                    h_map = {h: i for i, h in enumerate(headers)}
                    vegs = []
                    for r in rows[1:]:
                        if not r or all(c is None for c in r): continue
                        inst_id = clean_int(r[h_map.get('inst_id', 0)])
                        qty = clean_float(r[h_map.get('quantity', 6)])
                        v_code = str(r[h_map.get('v_code', 4)] or '').strip()
                        if not inst_id or not qty or not v_code: continue
                        purch_on = format_date(r[h_map.get('purchase_on', 2)])
                        purch_don = str(r[h_map.get('purchased_donation', 3)] or 'Purchase').strip()
                        bill_date = format_date(r[h_map.get('bill_date', 8)])
                        bill_no = str(r[h_map.get('bill_no', 9)] or '').strip()
                        rate = clean_float(r[h_map.get('rate', 10)])
                        rem = str(r[h_map.get('remarks', 13)] or '').strip()
                        issue_place = str(r[h_map.get('issue_place', 12)] or '').strip()
                        vegs.append((
                            inst_id, year_prefix, purch_on, v_code, qty, bill_date, bill_no, rate, rem, issue_place, purch_don
                        ))
                    if vegs:
                        q = '''
                            INSERT INTO Vegetable (Inst_ID, Year1, Purchase_On, V_Code, Quantity, Bill_Date, Bill_No, Rate, Remarks, Issue_Place, Purchased_Donation)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        ''' if db_type == "postgresql" else '''
                            INSERT INTO Vegetable (Inst_ID, Year1, Purchase_On, V_Code, Quantity, Bill_Date, Bill_No, Rate, Remarks, Issue_Place, Purchased_Donation)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?)
                        '''
                        chunk_size = 1000
                        for idx in range(0, len(vegs), chunk_size):
                            cursor.executemany(q, vegs[idx : idx + chunk_size])
                            conn.commit()
                            
        conn.close()
        print("Excel Register files loaded successfully into the database!")
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"Error seeding registers: {e}")

if __name__ == '__main__':
    init_db()
    app.run(debug=True, host='0.0.0.0', port=5000)
