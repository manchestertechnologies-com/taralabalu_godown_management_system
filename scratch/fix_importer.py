import os, sys, datetime, re

new_func = """def check_and_load_registers():
    try:
        # Check if already loaded
        existing = db_query("SELECT COUNT(*) as cnt FROM Stock_Issue")
        count = existing[0]['cnt'] if existing else 0
        if count >= 100:
            print(f"Historical registers already loaded ({count} issues found). Skipping initial import.")
            return
            
        print("Stock_Issue table is empty. Loading all historical Excel registers into database...")
        conn, db_type = get_db_connection()
        cursor = conn.cursor()
        
        # Clear any partial loads to ensure clean seeding
        cursor.execute("DELETE FROM Shops_Donors")
        cursor.execute("DELETE FROM Bills")
        cursor.execute("DELETE FROM Indents")
        cursor.execute("DELETE FROM Stock_Issue")
        cursor.execute("DELETE FROM Vegetable")
        conn.commit()

        base_dir = os.path.dirname(os.path.abspath(__file__))
        files_map = {
            '2025': 'Godown Register 2025-26.xlsx',
            '2026': 'Godown Register 2026-27.xlsx'
        }

        placeholder_donor_id = 999999
        q_donor_init = "INSERT INTO Shops_Donors (Year1, Shop_Donor_ID, Shop_Donor_Name, DateStamp) VALUES (%s,%s,%s,%s) ON CONFLICT DO NOTHING" if db_type == "postgresql" else "INSERT OR IGNORE INTO Shops_Donors (Year1, Shop_Donor_ID, Shop_Donor_Name, DateStamp) VALUES (?,?,?,?)"
        cursor.execute(q_donor_init, ('2025', placeholder_donor_id, 'Fallback General Donor', '2026-07-06'))
        conn.commit()

        # Define inline helper functions
        def format_date(val):
            if val is None: return ""
            import datetime
            if isinstance(val, (datetime.datetime, datetime.date)): return val.strftime('%Y-%m-%d')
            s = str(val).strip()
            if ' ' in s: s = s.split(' ')[0]
            return s

        def clean_float(val, default=0.0):
            if val is None: return default
            try: return float(val)
            except: return default

        def clean_int(val, default=0):
            if val is None: return default
            try: return int(float(val))
            except: return default

        import openpyxl
        import datetime

        for year_prefix, filename in files_map.items():
            filepath = os.path.join(base_dir, filename)
            if not os.path.exists(filepath):
                print(f"Excel Register file {filename} not found, skipping...")
                continue
                
            print(f"Importing {filename} for year {year_prefix}...")
            wb = openpyxl.load_workbook(filepath, data_only=True)
            
            # 1. Shops_Donors
            if 'Shops_Donors' in wb.sheetnames:
                ws = wb['Shops_Donors']
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) > 1:
                    headers = [str(h).strip().lower() for h in rows[0]]
                    h_map = {h: i for i, h in enumerate(headers)}
                    donors = []
                    for r in rows[1:]:
                        if not r or all(c is None for c in r): continue
                        did = clean_int(r[h_map.get('shop_donor_id', 0)])
                        name = str(r[h_map.get('name', 1)] or '').strip()
                        if not did or not name: continue
                        pid = clean_int(r[h_map.get('place_id')]) if 'place_id' in h_map else 0
                        place = str(r[h_map.get('place')] or '') if 'place' in h_map else ''
                        taluk = str(r[h_map.get('taluk')] or '') if 'taluk' in h_map else ''
                        dist = str(r[h_map.get('dist')] or '') if 'dist' in h_map else ''
                        pin = str(r[h_map.get('pin')] or '') if 'pin' in h_map else ''
                        mob = str(r[h_map.get('mobile')] or '') if 'mobile' in h_map else ''
                        rem = str(r[h_map.get('remarks')] or '') if 'remarks' in h_map else ''
                        donors.append((
                            year_prefix, did, name, pid, place, place, taluk, taluk, dist, dist, pin, 'Karnataka', 'India', mob, rem, datetime.date.today().strftime('%Y-%m-%d')
                        ))
                    if donors:
                        q = '''
                            INSERT INTO Shops_Donors 
                            (Year1, Shop_Donor_ID, Shop_Donor_Name, Place_ID, Place, Place_Kan, Taluk, Taluk_Kan, Dist, Dist_Kan, Pin, State, Country, Mobile, Remarks, DateStamp)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) ON CONFLICT (Shop_Donor_ID) DO NOTHING
                        ''' if db_type == "postgresql" else '''
                            INSERT OR IGNORE INTO Shops_Donors 
                            (Year1, Shop_Donor_ID, Shop_Donor_Name, Place_ID, Place, Place_Kan, Taluk, Taluk_Kan, Dist, Dist_Kan, Pin, State, Country, Mobile, Remarks, DateStamp)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                        '''
                        cursor.executemany(q, donors)
                        conn.commit()

            # 2. Bills
            if 'Bills' in wb.sheetnames:
                ws = wb['Bills']
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) > 1:
                    headers = [str(h).strip().lower() for h in rows[0]]
                    h_map = {h: i for i, h in enumerate(headers)}
                    bills = []
                    for r in rows[1:]:
                        if not r or all(c is None for c in r): continue
                        shop_id = clean_int(r[h_map.get('shop_id', 0)])
                        bill_date = format_date(r[h_map.get('bill_date', 2)])
                        bill_no = str(r[h_map.get('bill_no', 3)] or '').strip()
                        bill_amt = clean_float(r[h_map.get('bill_amount', 4)])
                        if not shop_id or not bill_date or not bill_no: continue
                        paid_by = str(r[h_map.get('paid_by')] or '') if 'paid_by' in h_map else ''
                        ch_date = format_date(r[h_map.get('ch_date')]) if 'ch_date' in h_map else ''
                        ch_no = str(r[h_map.get('ch_no') or '']) if 'ch_no' in h_map else ''
                        ch_amt = clean_float(r[h_map.get('ch_amount')]) if 'ch_amount' in h_map else 0.0
                        rem = str(r[h_map.get('remarks')] or '') if 'remarks' in h_map else ''
                        bills.append((
                            year_prefix, shop_id, bill_date, bill_no, bill_amt, paid_by, ch_date, ch_no, ch_amt, rem, datetime.date.today().strftime('%Y-%m-%d')
                        ))
                    if bills:
                        q = '''
                            INSERT INTO Bills (Year1, Shop_Donor_ID, Bill_Date, Bill_No, Bill_Amount, Paid_By, Ch_Date, Ch_No, Ch_Amount, Remarks, DateAdded)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        ''' if db_type == "postgresql" else '''
                            INSERT INTO Bills (Year1, Shop_Donor_ID, Bill_Date, Bill_No, Bill_Amount, Paid_By, Ch_Date, Ch_No, Ch_Amount, Remarks, DateAdded)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?)
                        '''
                        cursor.executemany(q, bills)
                        conn.commit()

            # 3. Indents
            if 'Indents' in wb.sheetnames:
                ws = wb['Indents']
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) > 1:
                    headers = [str(h).strip().lower().replace(" ", "_") for h in rows[0]]
                    h_map = {h: i for i, h in enumerate(headers)}
                    indents = []
                    for r in rows[1:]:
                        if not r or all(c is None for c in r): continue
                        code = clean_int(r[h_map.get('code', 3)])
                        inst_id = clean_int(r[h_map.get('inst_id', 1)])
                        qty = clean_float(r[h_map.get('quantity', 5)])
                        indent_no = str(r[h_map.get('indent_no', 9)] or '').strip()
                        if not code or not inst_id or not qty or not indent_no: continue
                        indent_date = format_date(r[h_map.get('indent_date', 0)])
                        sanc_qty = clean_float(r[h_map.get('sanctioned_quantity', 10)])
                        status = str(r[h_map.get('sanctioned', 11)] or 'Pending').strip()
                        sanc_on = format_date(r[h_map.get('sanctioned_on')]) if 'sanctioned_on' in h_map else ''
                        indents.append((
                            year_prefix, indent_date, code, inst_id, qty, indent_no, sanc_qty, status, sanc_on, datetime.date.today().strftime('%Y-%m-%d'), 'Imported'
                        ))
                    if indents:
                        q = '''
                            INSERT INTO Indents (Year1, Indent_Date, Grocery_Code, Inst_ID, Quantity, Indent_no, Sanctioned_Quantity, Sanctioned, Sanctioned_on, DateStamp, Remarks)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        ''' if db_type == "postgresql" else '''
                            INSERT INTO Indents (Year1, Indent_Date, Grocery_Code, Inst_ID, Quantity, Indent_no, Sanctioned_Quantity, Sanctioned, Sanctioned_on, DateStamp, Remarks)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?)
                        '''
                        cursor.executemany(q, indents)
                        conn.commit()
                        
            # 3.5 Godown (Opening Stock)
            if 'Godown' in wb.sheetnames:
                ws = wb['Godown']
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) > 1:
                    headers = [str(h).strip().lower().replace(" ", "_") for h in rows[0]]
                    h_map = {h: i for i, h in enumerate(headers)}
                    openings = []
                    opening_date = f"{year_prefix}-04-01"
                    for r in rows[1:]:
                        if not r or all(c is None for c in r): continue
                        code = clean_int(r[h_map.get('grocery_code', 1)])
                        open_qty = clean_float(r[h_map.get('opening_stock', 4)])
                        if not code or open_qty <= 0: continue
                        rate = clean_float(r[h_map.get('std_rate', 9)])
                        amt = open_qty * rate
                        openings.append((
                            year_prefix, opening_date, code, placeholder_donor_id, '', '', '', '', '', rate, amt, 'Yes', open_qty, None, 0.0, 0.0, 'Central Godown', datetime.date.today().strftime('%Y-%m-%d'), 'Imported Opening Stock', 'Opening Stock'
                        ))
                    if openings:
                        q = '''
                            INSERT INTO Stock_Issue (Year1, Date1, Grocery_Code, Shop_Donor_ID, Book_No, Receipt_No, Receipt_Date, Bill_Date, Bill_No, Purchase_Rate, Purchase_Amount, Paid, Stock, Issue_Inst_ID, Issue, Issue_Amount, Received_By, DateStamp, Remarks, Purchased_Donation)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        ''' if db_type == "postgresql" else '''
                            INSERT INTO Stock_Issue (Year1, Date1, Grocery_Code, Shop_Donor_ID, Book_No, Receipt_No, Receipt_Date, Bill_Date, Bill_No, Purchase_Rate, Purchase_Amount, Paid, Stock, Issue_Inst_ID, Issue, Issue_Amount, Received_By, DateStamp, Remarks, Purchased_Donation)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                        '''
                        chunk_size = 1000
                        for idx in range(0, len(openings), chunk_size):
                            cursor.executemany(q, openings[idx : idx + chunk_size])
                            conn.commit()

            # 4. Stock (Inwards)
            if 'Stock' in wb.sheetnames:
                ws = wb['Stock']
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) > 1:
                    headers = [str(h).strip().lower().replace(" ", "_") for h in rows[0]]
                    h_map = {h: i for i, h in enumerate(headers)}
                    inwards = []
                    for r in rows[1:]:
                        if not r or all(c is None for c in r): continue
                        code = clean_int(r[h_map.get('grocery_code', 3)])
                        stock_qty = clean_float(r[h_map.get('stock', 5)])
                        if not code or not stock_qty: continue
                        date1 = format_date(r[h_map.get('date', 1)])
                        purch_don = str(r[h_map.get('purchased_donation', 2)] or 'Purchase').strip()
                        rate = clean_float(r[h_map.get('purchase_rate', 7)])
                        amt = clean_float(r[h_map.get('purchase_amount', 9)])
                        book_no = str(r[h_map.get('book_no')] or '') if 'book_no' in h_map else ''
                        receipt_no = str(r[h_map.get('receipt_no')] or '') if 'receipt_no' in h_map else ''
                        receipt_date = format_date(r[h_map.get('receipt_date')]) if 'receipt_date' in h_map else ''
                        
                        donor_id = clean_int(r[h_map.get('shop_donor_id')]) if 'shop_donor_id' in h_map else placeholder_donor_id
                        if not donor_id: donor_id = placeholder_donor_id
                        bill_date = format_date(r[h_map.get('bill_date')]) if 'bill_date' in h_map else ''
                        bill_no = str(r[h_map.get('bill_no')] or '') if 'bill_no' in h_map else ''
                        paid = str(r[h_map.get('paid')] or 'No') if 'paid' in h_map else 'No'
                        received_by = str(r[h_map.get('received_by')] or 'Central Godown') if 'received_by' in h_map else 'Central Godown'
                        
                        inwards.append((
                            year_prefix, date1, code, donor_id, book_no, receipt_no, receipt_date, bill_date, bill_no, rate, amt, paid, stock_qty, None, 0.0, 0.0, received_by, datetime.date.today().strftime('%Y-%m-%d'), 'Imported', purch_don
                        ))
                    if inwards:
                        q = '''
                            INSERT INTO Stock_Issue (Year1, Date1, Grocery_Code, Shop_Donor_ID, Book_No, Receipt_No, Receipt_Date, Bill_Date, Bill_No, Purchase_Rate, Purchase_Amount, Paid, Stock, Issue_Inst_ID, Issue, Issue_Amount, Received_By, DateStamp, Remarks, Purchased_Donation)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        ''' if db_type == "postgresql" else '''
                            INSERT INTO Stock_Issue (Year1, Date1, Grocery_Code, Shop_Donor_ID, Book_No, Receipt_No, Receipt_Date, Bill_Date, Bill_No, Purchase_Rate, Purchase_Amount, Paid, Stock, Issue_Inst_ID, Issue, Issue_Amount, Received_By, DateStamp, Remarks, Purchased_Donation)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                        '''
                        chunk_size = 1000
                        for idx in range(0, len(inwards), chunk_size):
                            cursor.executemany(q, inwards[idx : idx + chunk_size])
                            conn.commit()

            # 5. Issue (Outwards)
            if 'Issue' in wb.sheetnames:
                ws = wb['Issue']
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) > 1:
                    headers = [str(h).strip().lower().replace(" ", "_") for h in rows[0]]
                    h_map = {h: i for i, h in enumerate(headers)}
                    outwards = []
                    for r in rows[1:]:
                        if not r or all(c is None for c in r): continue
                        code = clean_int(r[h_map.get('code', 4)])
                        issue_qty = clean_float(r[h_map.get('issue', 6)])
                        inst_id = clean_int(r[h_map.get('issue_inst_id', 1)])
                        if not code or not issue_qty or not inst_id: continue
                        date1 = format_date(r[h_map.get('date', 3)])
                        rate = clean_float(r[h_map.get('std_rate', 8)])
                        amt = clean_float(r[h_map.get('issue_amt', 9)])
                        outwards.append((
                            year_prefix, date1, code, None, '', '', '', '', '', rate, 0.0, 'No', 0.0, inst_id, issue_qty, amt, 'Receiver', datetime.date.today().strftime('%Y-%m-%d'), 'Imported Outward', 'Issue'
                        ))
                    if outwards:
                        q = '''
                            INSERT INTO Stock_Issue (Year1, Date1, Grocery_Code, Shop_Donor_ID, Book_No, Receipt_No, Receipt_Date, Bill_Date, Bill_No, Purchase_Rate, Purchase_Amount, Paid, Stock, Issue_Inst_ID, Issue, Issue_Amount, Received_By, DateStamp, Remarks, Purchased_Donation)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        ''' if db_type == "postgresql" else '''
                            INSERT INTO Stock_Issue (Year1, Date1, Grocery_Code, Shop_Donor_ID, Book_No, Receipt_No, Receipt_Date, Bill_Date, Bill_No, Purchase_Rate, Purchase_Amount, Paid, Stock, Issue_Inst_ID, Issue, Issue_Amount, Received_By, DateStamp, Remarks, Purchased_Donation)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                        '''
                        chunk_size = 1000
                        for idx in range(0, len(outwards), chunk_size):
                            cursor.executemany(q, outwards[idx : idx + chunk_size])
                            conn.commit()

            # 6. Vegetable
            if 'Vegetable' in wb.sheetnames:
                ws = wb['Vegetable']
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) > 1:
                    headers = [str(h).strip().lower().replace(" ", "_") for h in rows[0]]
                    h_map = {h: i for i, h in enumerate(headers)}
                    vegs = []
                    for r in rows[1:]:
                        if not r or all(c is None for c in r): continue
                        inst_id = clean_int(r[h_map.get('inst_id', 0)])
                        qty = clean_float(r[h_map.get('quantity', 6)])
                        v_code = str(r[h_map.get('v_code', 4)] or '').strip()
                        if not inst_id or not qty or not v_code: continue
                        purch_on = format_date(r[h_map.get('purchase_on', 2)])
                        purch_don = str(r[h_map.get('purchased_donation', 3)] or 'Purchase').strip()
                        bill_date = format_date(r[h_map.get('bill_date', 8)])
                        bill_no = str(r[h_map.get('bill_no', 9)] or '').strip()
                        rate = clean_float(r[h_map.get('rate', 10)])
                        rem = str(r[h_map.get('remarks', 13)] or '').strip()
                        issue_place = str(r[h_map.get('issue_place', 12)] or '').strip()
                        vegs.append((
                            inst_id, year_prefix, purch_on, v_code, qty, bill_date, bill_no, rate, rem, issue_place, purch_don
                        ))
                    if vegs:
                        q = '''
                            INSERT INTO Vegetable (Inst_ID, Year1, Purchase_On, V_Code, Quantity, Bill_Date, Bill_No, Rate, Remarks, Issue_Place, Purchased_Donation)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        ''' if db_type == "postgresql" else '''
                            INSERT INTO Vegetable (Inst_ID, Year1, Purchase_On, V_Code, Quantity, Bill_Date, Bill_No, Rate, Remarks, Issue_Place, Purchased_Donation)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?)
                        '''
                        chunk_size = 1000
                        for idx in range(0, len(vegs), chunk_size):
                            cursor.executemany(q, vegs[idx : idx + chunk_size])
                            conn.commit()
                            
        conn.close()
        print("Excel Register files loaded successfully into the database!")
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"Error seeding registers: {e}")"""

with open('app.py', 'r', encoding='utf-8') as f:
    content = f.read()

import re
# Find the start of check_and_load_registers
start_idx = content.find('def check_and_load_registers():')
if start_idx == -1:
    print("Error: check_and_load_registers not found")
    sys.exit(1)

# Find the start of the next function
end_idx = content.find('if __name__ ==', start_idx)
if end_idx == -1:
    print("Error: __name__ == '__main__' not found")
    sys.exit(1)

new_content = content[:start_idx] + new_func + '\\n\\n' + content[end_idx:]

with open('app.py', 'w', encoding='utf-8') as f:
    f.write(new_content)

print("Successfully replaced check_and_load_registers in app.py")
